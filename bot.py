import asyncio
import logging
import os
import json
import re
import tempfile
from typing import List, Optional, Tuple, Dict
from datetime import datetime, timedelta, timezone

# --- корректная работа с часовым поясом МСК (Windows + Linux) ---
try:
    from zoneinfo import ZoneInfo
    MSK = ZoneInfo("Europe/Moscow")
except Exception:
    # fallback: UTC+3 без сезонных сдвигов (на крайний случай, если нет tzdata на Windows)
    MSK = timezone(timedelta(hours=3))

from pathlib import Path

from aiogram import Bot, Dispatcher, F
from aiogram.filters import CommandStart
from aiogram.fsm.context import FSMContext
from aiogram.fsm.state import State, StatesGroup
from aiogram.types import CallbackQuery, FSInputFile, Message
from aiogram.utils.keyboard import InlineKeyboardBuilder
from aiogram.exceptions import TelegramBadRequest
from sqlalchemy import desc, select

from config import BOT_TOKEN, MAX_BLUE, is_admin, ENABLE_ADMIN_CREATE_PLAYER
from db import (
    Player,
    Game,
    GameParticipant,
    Session,
    create_game,
    create_player,
    delete_game,
    delete_player_if_no_games,
    get_game,
    init_db,
    list_all_games,
    update_player_name,
    create_purchase,
    list_purchases,
    set_purchase_received,
)

from services import (
    apply_ratings,
    get_team_rosters,
    recompute_all_galleons,
    recompute_all_ratings,
    search_players,
    set_result_type_and_killer,
    set_voldemort,
    set_team_roster,
    validate_rosters,
    get_player_streaks,
)

logging.basicConfig(level=logging.INFO)
bot = Bot(BOT_TOKEN)
dp = Dispatcher()

# ===================== Patches / helpers =====================
# Safer answer for old callback queries (ignores 'query is too old' / invalid ID)
async def safe_answer(c: CallbackQuery, *args, **kwargs):
    try:
        return await c.answer(*args, **kwargs)
    except TelegramBadRequest as e:
        msg = str(e).lower()
        if 'query is too old' in msg or 'query id is invalid' in msg or 'timeout expired' in msg:
            return
        # игнорируем "Bad Request: message can't be edited" и подобные не критичные
        if 'bad request' in msg:
            return
        raise
    except Exception:
        return

# Make admin check tolerant to case and '@' and fallback to ADMIN_USERNAMES from .env
try:
    _is_admin_base = is_admin  # from config
except Exception:
    async def _is_admin_base(*_args, **_kwargs):
        return False

def is_admin(user_id: int, username: Optional[str]) -> bool:  # type: ignore[override]
    try:
        if _is_admin_base(user_id, username):  # if config says admin — trust it
            return True
    except Exception:
        pass
    uname = (username or "").strip().lstrip("@").lower()
    env_val = os.getenv("ADMIN_USERNAMES", "") or ""
    # allow comma/semicolon separated values with/without '@' and arbitrary spaces
    env_names = [x.strip().lstrip("@").lower() for x in re.split(r"[;,]", env_val) if x.strip()]
    return bool(uname and uname in env_names)


# ===================== Shop / Galleons =====================
COIN = "💰"

SHOP_ITEMS = [
    {"code": "pm_first_game", "label": "Первый министр (1-я игра)", "title": "Заявиться первым министром в первой игре вечера (до раздачи ролей)", "cost": 5, "emoji": "👑"},
    {"code": "pm_replace_lord", "label": "Первый министр (смещение)", "title": "Заявиться первым министром сместив прошлого лорда", "cost": 15, "emoji": "🛡️"},
    {"code": "badge", "label": "Фирменный значок", "title": "Фирменный значок", "cost": 100, "emoji": "🏷️"},
    {"code": "random_12_rooms", "label": "Сертификат 12 комнат", "title": "Случайный сертификат 12 комнат", "cost": 300, "emoji": "🎟️"},
    {"code": "named_ballot", "label": "Именная голосовалка", "title": "Именная голосовалка", "cost": 300, "emoji": "🗳️"},
]

def _msk_now_str() -> str:
    return datetime.now(MSK).strftime("%d.%m.%Y %H:%M:%S (МСК)")

def shop_menu_kb():
    kb = InlineKeyboardBuilder()
    for item in SHOP_ITEMS:
        kb.button(text=f"{item['emoji']} {item.get('label', item['title'])} — {item['cost']}{COIN}", callback_data=f"shop:buy:{item['code']}")
    kb.button(text="⬅️ Назад", callback_data="backhome")
    kb.adjust(1)
    return kb.as_markup()

def mypurchases_list_kb(purchases: list):
    kb = InlineKeyboardBuilder()
    for p in purchases:
        mark = "✅" if p.is_received else "❌"
        kb.button(text=f"{mark} {p.title} — {p.cost}{COIN} • {p.created_at.strftime('%d.%m %H:%M')} ", callback_data=f"mypur:item:{p.id}")
    kb.button(text="⬅️ Назад", callback_data="backhome")
    kb.adjust(1)
    return kb.as_markup()

def purchase_status_kb(purchase_id: int):
    kb = InlineKeyboardBuilder()
    kb.button(text="Получено ✅", callback_data=f"mypur:set:{purchase_id}:1")
    kb.button(text="Не получено ❌", callback_data=f"mypur:set:{purchase_id}:0")
    kb.button(text="⬅️ Назад", callback_data="mypur:menu")
    kb.adjust(1)
    return kb.as_markup()


# ===================== small helpers =====================
# ---- Guard: warn when leaving unfinished game creation ----
def _encode_target(s: str) -> str:
    return s.replace(":", "§")

def _decode_target(s: str) -> str:
    return s.replace("§", ":")

def confirm_leave_kb(gid: int, target: str):
    kb = InlineKeyboardBuilder()
    kb.button(text="Да, выйти", callback_data=f"leave:confirm:{gid}:{_encode_target(target)}")
    kb.button(text="Нет, остаться", callback_data=f"leave:stay:{gid}")
    kb.adjust(2)
    return kb.as_markup()

async def _maybe_warn_unfinished(c: CallbackQuery, state: FSMContext, target: str) -> bool:
    data = await state.get_data()
    gid = data.get("pending_gid")
    if not gid:
        return False
    async with Session() as session:
        g = await get_game(session, gid)
    if not g:
        await state.update_data(pending_gid=None)
        return False
    if getattr(g, "result_type", None):
        await state.update_data(pending_gid=None)
        return False
    txt = "Состав команд не заполнен — при выходе он будет сброшен и игра не будет записана.\nПерейти в другой раздел?"
    await safe_edit(c.message, txt, reply_markup=confirm_leave_kb(gid, target))
    await safe_answer(c, )
    return True

def full_name(p: Player) -> str:
    return f"{p.first_name}{(' ' + p.last_name) if p.last_name else ''}"

def now_msk() -> datetime:
    return datetime.now(MSK)

async def safe_edit(message, text, **kwargs):
    same_text = (message.text or message.caption or "") == (text or "")
    same_markup = "reply_markup" in kwargs and getattr(message, "reply_markup", None) == kwargs["reply_markup"]
    if same_text and ("reply_markup" not in kwargs or same_markup):
        return message
    try:
        return await message.edit_text(text, **kwargs)
    except TelegramBadRequest as e:
        if "message is not modified" in str(e).lower():
            return message
        raise


def _strip_repeat_summary(summary: str) -> str:
    """
    Удаляет дублирующийся блок "Игра завершена./Победа .../Средний MMR .../Фаворит матча ..."
    из текста, возвращаемого apply_ratings(), чтобы не было повтора в финальном сообщении.
    """
    lines = []
    skip = False
    for raw in (summary or "").splitlines():
        s = raw.strip()
        if s.startswith("Игра завершена."):
            skip = True
            continue
        if skip and (s.startswith("Победа ") or s.startswith("Средний MMR") or s.startswith("Фаворит матча")):
            continue
        if skip and (s.startswith("Изменение MMR")):
            skip = False
        lines.append(raw)
    text = "\n".join(lines).strip()
    return text

def _normalize_summary_delta(summary: str) -> str:
    """Оставляет из summary только строку с дельтой MMR и
    переименовывает Синие/Красные в Орден/Пожиратели.
    """
    if not summary:
        return ""
    text = summary.replace("Синие", "Орден").replace("Красные", "Пожиратели")
    i = text.find("Изменение MMR")
    return text[i:].strip() if i != -1 else text.strip()
def roster_block(title: str, players: List[Player], vold: Optional[Player]) -> str:
    def line(p: Player) -> str:
        tag = " (Воланд)" if (vold and p.id == vold.id) else ""
        return f"- {full_name(p)} [{p.rating}]{tag}"
    body = "\n".join(line(p) for p in players) if players else "—"
    return f"{title} ({len(players)}):\n{body}"


async def roster_summary(session: Session, game_id: int) -> Tuple[str, List[Player], List[Player], Optional[Player]]:
    blue, red, vold = await get_team_rosters(session, game_id)
    ok, msg = await validate_rosters(blue, red, vold)
    blue_block = roster_block('🟦 Орден Феникса', blue, vold)
    red_block = roster_block('🟪 Пожиратели + Воландеморт', red, vold)
    # Показываем Воландеморта отдельно в красном блоке,
    # если он выбран, но не находится в списке red.
    if vold and all(p.id != vold.id for p in red):
        suffix = f"- {full_name(vold)} [{vold.rating}] (Воланд)"
        # red_block имеет вид: "Заголовок\nТело"
        parts = red_block.split("\n", 1)
        title = parts[0]
        body = parts[1] if len(parts) > 1 else "—"
        body = suffix if body.strip() == "—" else body + "\n" + suffix
        red_block = title + "\n" + body
    text = f"{blue_block}\n\n{red_block}\n\nСтатус: {('✅' if ok else '❌')} {msg}"
    return text, blue, red, vold

RESULT_HUMAN = {
    "blue_laws": "Победа Ордена Феникса — выложены 5 синих законов",
    "blue_kill": "Победа Ордена Феникса — Воландеморт отправлен в Азкабан",
    "red_laws": "Победа Пожирателей — выложены 6 красных законов",
    "red_director": "Победа Пожирателей — Воландеморт избран директором",
}

def favorite_side(blue_avg: float, red_avg: float) -> str:
    if abs(blue_avg - red_avg) < 1e-9:
        return "неопределён (средние равны)"
    return "Орден Феникса" if blue_avg > red_avg else "Пожиратели"

# ===================== persist (day list, apps, auth, notes, metrics) =====================
DAY_LIST_PATH = Path("day_list.json")
APPS_PATH = Path("applications.json")
AUTH_MAP_PATH = Path("auth_map.json")
NOTES_PATH = Path("game_notes.json")
METRICS_PATH = Path("bot_metrics.json")  # счётчики бота (без нагрузки на БД)

def _load_json_list(path: Path) -> list:
    if not path.exists():
        return []
    try:
        with path.open("r", encoding="utf-8") as f:
            return json.load(f) or []
    except Exception:
        return []

def _save_json_list(path: Path, data: list) -> None:
    with path.open("w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False)

def _load_json_obj(path: Path) -> dict:
    if not path.exists():
        return {}
    try:
        with path.open("r", encoding="utf-8") as f:
            return json.load(f) or {}
    except Exception:
        return {}

def _save_json_obj(path: Path, data: dict) -> None:
    with path.open("w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False)

# ---- day list
def _load_day_list() -> List[int]:
    # авто-создание и защита на случай битого содержимого
    if not DAY_LIST_PATH.exists():
        _save_json_list(DAY_LIST_PATH, [])
        return []
    try:
        data = _load_json_list(DAY_LIST_PATH)
        if not isinstance(data, list):
            data = []
    except Exception:
        data = []
    return data
def _save_day_list(ids: List[int]) -> None:
    _save_json_list(DAY_LIST_PATH, ids)

# ---- applications
def _load_apps() -> List[dict]:
    return _load_json_list(APPS_PATH)
def _save_apps(apps: List[dict]) -> None:
    _save_json_list(APPS_PATH, apps)

# ---- auth map
def _load_auth_map() -> Dict[str, int]:
    return _load_json_obj(AUTH_MAP_PATH)
def _save_auth_map(mapping: Dict[str, int]) -> None:
    _save_json_obj(AUTH_MAP_PATH, mapping)
def is_authorized_user(user_id: int) -> bool:
    return str(user_id) in _load_auth_map()
def get_player_id_for_user(user_id: int) -> Optional[int]:
    return _load_auth_map().get(str(user_id))
def link_user_to_player(user_id: int, player_id: int) -> None:
    mp = _load_auth_map()
    mp[str(user_id)] = player_id
    _save_auth_map(mp)

# ---- game notes
def _get_notes(game_id: int) -> List[dict]:
    obj = _load_json_obj(NOTES_PATH)
    return obj.get(str(game_id), [])
def _add_note(game_id: int, user_id: int, text: str):
    obj = _load_json_obj(NOTES_PATH)
    obj.setdefault(str(game_id), []).append({
        "ts": now_msk().isoformat(),
        "user_id": user_id,
        "text": text.strip()
    })
    _save_json_obj(NOTES_PATH, obj)
def _has_notes(game_id: int) -> bool:
    obj = _load_json_obj(NOTES_PATH)
    return bool(obj.get(str(game_id)))

# ---- bot metrics (лёгкие счётчики и множества уникальных пользователей по датам)
def _metrics() -> dict:
    m = _load_json_obj(METRICS_PATH)
    if not m:
        m = {
            "counters": {
                "games_created": 0,
                "games_finished": 0,
                "excel_downloads": 0,
                "auth_approved": 0,
                "visits": 0,
            },
            "by_day": {},  # "YYYY-MM-DD": {"active_user_ids": [..], "clicks": N}
        }
    return m

def _save_metrics(m: dict):
    _save_json_obj(METRICS_PATH, m)

def metric_visit(user_id: int):
    m = _metrics()
    m["counters"]["visits"] += 1
    day = now_msk().date().isoformat()
    m["by_day"].setdefault(day, {"active_user_ids": [], "clicks": 0})
    if user_id not in m["by_day"][day]["active_user_ids"]:
        m["by_day"][day]["active_user_ids"].append(user_id)
    _save_metrics(m)

def metric_click(user_id: int, weight: int = 1):
    m = _metrics()
    day = now_msk().date().isoformat()
    m["by_day"].setdefault(day, {"active_user_ids": [], "clicks": 0})
    m["by_day"][day]["clicks"] += int(weight)
    if user_id not in m["by_day"][day]["active_user_ids"]:
        m["by_day"][day]["active_user_ids"].append(user_id)
    _save_metrics(m)

def metric_inc(key: str):
    m = _metrics()
    m["counters"][key] = int(m["counters"].get(key, 0)) + 1
    _save_metrics(m)


def _metrics_summary(mode: str) -> tuple[str, dict]:
    """
    mode: 'week' | 'month' | 'all'
    Returns (text, stats_dict)
    """
    m = _metrics()
    by_day = m.get("by_day", {}) or {}
    # Determine date range
    today = now_msk().date()
    if mode == "week":
        cutoff = today - timedelta(days=7)
        title = "за неделю"
    elif mode == "month":
        cutoff = today - timedelta(days=30)
        title = "за месяц"
    else:
        cutoff = None
        title = "за всё время"

    clicks = 0
    active_users = set()
    days_considered = 0

    for day_str, rec in sorted(by_day.items()):
        try:
            day_date = datetime.fromisoformat(day_str).date()
        except Exception:
            continue
        if cutoff and day_date < cutoff:
            continue
        days_considered += 1
        clicks += int(rec.get("clicks", 0) or 0)
        for uid in rec.get("active_user_ids", []) or []:
            try:
                active_users.add(int(uid))
            except Exception:
                pass

    counters = m.get("counters", {})
    total_games = int(counters.get("games_created", 0) or 0)
    total_finished = int(counters.get("games_finished", 0) or 0)
    excel_downloads = int(counters.get("excel_downloads", 0) or 0)
    auth_approved = int(counters.get("auth_approved", 0) or 0)

    text = (
        f"<b>Статистика бота {title}</b>\n\n"
        f"• Активных пользователей: <b>{len(active_users)}</b>\n"
        f"• Клики: <b>{clicks}</b>\n"
        f"• Дней в выборке: <b>{days_considered}</b>\n\n"
        f"<u>Счётчики всего времени</u>\n"
        f"• Создано игр: <b>{total_games}</b>\n"
        f"• Завершено игр: <b>{total_finished}</b>\n"
        f"• Выгрузок Excel: <b>{excel_downloads}</b>\n"
        f"• Одобрений авторизации: <b>{auth_approved}</b>"
    )

    stats = {
        "mode": mode,
        "active_users": len(active_users),
        "clicks": clicks,
        "days": days_considered,
        "totals": {
            "games_created": total_games,
            "games_finished": total_finished,
            "excel_downloads": excel_downloads,
            "auth_approved": auth_approved,
        },
    }
    return text, stats


# ===================== FSM =====================
class CreateGameFSM(StatesGroup):
    main_menu = State()
    selecting_team = State()
    search_player_for = State()
    wait_pick_killer = State()
    wait_note_text = State()

class AdminFSM(StatesGroup):
    wait_new_fullname = State()
    wait_new_player = State()

class UserAuthFSM(StatesGroup):
    wait_name = State()

# ===================== Keyboards =====================
def home_kb_for_user(is_admin_flag: bool, is_authorized: bool):
    kb = InlineKeyboardBuilder()
    if is_admin_flag:
        kb.button(text="➕ Добавить игру", callback_data="start:newgame")
    kb.button(text="📚 Завершённые игры", callback_data="finished:menu")
    kb.button(text="🧩 Игрок дня", callback_data="playeroftheday")
    kb.button(text="🏆 Рейтинг игроков", callback_data="rating:menu")
    if is_authorized:
        kb.button(text="📈 Моя статистика", callback_data="me:stats")
        kb.button(text=f"{COIN} Мои Галлеоны", callback_data="me:galleons")
        kb.button(text="🛒 Лавка Олливандера", callback_data="shop:menu")
        kb.button(text="🧾 Мои покупки", callback_data="mypur:menu")
    kb.button(text="❓ FAQ", callback_data="faq")
    if not is_authorized:
        if is_admin_flag:
            kb.button(text="👤 Авторизоваться (админ)", callback_data="auth:start")
        else:
            kb.button(text="🔐 Авторизация", callback_data="auth:start")
    if is_admin_flag:
        kb.button(text="🛠 Админ-панель", callback_data="admin:menu")
    kb.button(text="⬅️ В главное меню", callback_data="backhome")
    kb.adjust(1)
    return kb.as_markup()

def main_menu_kb(game_id: int):
    kb = InlineKeyboardBuilder()
    kb.button(text="Команда Ордена Феникса", callback_data=f"multiteam:blue:{game_id}")
    kb.button(text="Команда Пожирателей", callback_data=f"multiteam:red:{game_id}")
    kb.button(text="Кто был Воландемортом", callback_data=f"vold:{game_id}")
    kb.button(text="Проверить заполнение", callback_data=f"check:{game_id}")
    kb.button(text="Выберите победителя", callback_data=f"winner:{game_id}")
    kb.button(text="🏆 Рейтинг игроков", callback_data="rating:menu")
    kb.button(text="📚 Завершённые игры", callback_data="finished:menu")
    kb.button(text="⬅️ В главное меню", callback_data="backhome")
    kb.adjust(1)
    return kb.as_markup()

def after_finish_kb():
    kb = InlineKeyboardBuilder()
    kb.button(text="📚 Завершённые игры", callback_data="finished:menu")
    kb.button(text="🏆 Рейтинг игроков", callback_data="rating:menu")
    kb.button(text="⬅️ В главное меню", callback_data="backhome")
    kb.adjust(1)
    return kb.as_markup()

def rating_kb():
    kb = InlineKeyboardBuilder()
    kb.button(text="🌟 Лучшие синие", callback_data="rating:top:blue")
    kb.button(text="🔥 Лучшие пожиратели", callback_data="rating:top:red")
    kb.button(text="🕶️ Лучшие Воландеморты", callback_data="rating:top:vold")
    kb.button(text="🗡️ Лучшие киллеры", callback_data="rating:top:killer")
    kb.button(text="⬇️ Экспорт в Excel", callback_data="rating:export")
    kb.button(text="⬅️ В главное меню", callback_data="backhome")
    kb.adjust(1)
    return kb.as_markup()

def admin_menu_kb():
    pending = len([a for a in _load_apps() if a.get("status") == "pending"])
    inbox_text = "📫 Заявки в Бота" + (f" 🔴 ({pending})" if pending else "")
    kb = InlineKeyboardBuilder()
    kb.button(text="🧑‍🤝‍🧑 Игроки (редакт/удал.)", callback_data="admin:players")
    kb.button(text="🎮 Игры (удаление)", callback_data="admin:games")
    kb.button(text="🔁 Пересчитать рейтинг (все)", callback_data="admin:recompute")
    kb.button(text=f"{COIN} Перерасчет Галлеонов", callback_data="admin:recompute_galleons")
    kb.button(text="📋 Список дня", callback_data="admin:daylist")
    kb.button(text=inbox_text, callback_data="admin:apps")
    kb.button(text="📈 Статистика бота", callback_data="botstats:menu")
    kb.button(text="ℹ️ Инфо для админов", callback_data="admin:info")
    kb.button(text="⬅️ В главное меню", callback_data="backhome")
    kb.adjust(1)
    return kb.as_markup()

def admin_games_kb(games):
    kb = InlineKeyboardBuilder()
    for g in games[-50:]:
        # показываем время создания
        title = getattr(g, "title", "Игра")
        kb.button(text=f"🗑 ID {g.id}: {title}", callback_data=f"admin:game:del:{g.id}")
    kb.button(text="⬅️ Назад", callback_data="admin:menu")
    kb.adjust(1)
    return kb.as_markup()

def _status_prefix(pid: int, selected_ids: List[int], blue_ids: List[int], red_ids: List[int], vold_id: Optional[int], color_for_team: str) -> str:
    if pid in selected_ids:
        idx = selected_ids.index(pid) + 1
        return f"{color_for_team} #{idx} "
    if vold_id and pid == vold_id:
        return "🟣 "
    if pid in blue_ids:
        return "🔵 "
    if pid in red_ids:
        return "🔴 "
    return ""

def multiselect_kb(
    players: List[Player],
    selected_ids: List[int],
    team: str,
    game_id: int,
    limit: int,
    vold_id: Optional[int],
    admin_can_add: bool,
    blue_ids: List[int],
    red_ids: List[int],
):
    color = "🔵" if team == "blue" else "🔴"
    kb = InlineKeyboardBuilder()
    for p in players:
        prefix = _status_prefix(p.id, selected_ids, blue_ids, red_ids, vold_id, color)
        suffix = " (Воланд)" if vold_id and p.id == vold_id else ""
        kb.button(
            text=f"{prefix}{full_name(p)}{suffix} [{p.rating}]",
            callback_data=f"toggle:{team}:{game_id}:{p.id}",
        )
    kb.button(text="🔎 Поиск", callback_data=f"search:{team}:{game_id}")
    # ⛔️ больше НЕ создаём игрока из набора команд — только через авторизацию
    kb.button(text="🧹 Очистить выбор", callback_data=f"clear:{team}:{game_id}")
    kb.button(text="✅ Сохранить команду", callback_data=f"save:{team}:{game_id}")
    kb.button(text="⬅️ Назад", callback_data=f"back:{game_id}")
    kb.adjust(1)
    return kb.as_markup()

def source_choice_kb(team: str, game_id: int):
    kb = InlineKeyboardBuilder()
    kb.button(text="📋 Список дня", callback_data=f"source:day:{team}:{game_id}")
    kb.button(text="🗂 Все игроки", callback_data=f"source:all:{team}:{game_id}")
    kb.button(text="⬅️ Назад", callback_data=f"back:{game_id}")
    kb.adjust(1)
    return kb.as_markup()

def daylist_kb(all_players: List[Player], ids: List[int]):
    chosen = set(ids)
    kb = InlineKeyboardBuilder()
    for p in [p for p in all_players if p.id in chosen]:
        kb.button(text=f"✅ {full_name(p)} (ID {p.id})", callback_data=f"day:toggle:{p.id}")
    for p in [p for p in all_players if p.id not in chosen]:
        kb.button(text=f"{full_name(p)} (ID {p.id})", callback_data=f"day:toggle:{p.id}")
    kb.button(text="💾 Сохранить список", callback_data="day:save")
    kb.button(text="🧹 Очистить список", callback_data="day:clear")
    kb.button(text="🗑 Удалить выборочно", callback_data="day:mode:del")
    kb.button(text="⬅️ Назад", callback_data="admin:menu")
    kb.adjust(1)
    return kb.as_markup()

@dp.callback_query(F.data == "admin:daylist")
async def admin_daylist(c: CallbackQuery):
    metric_click(c.from_user.id)
    if not is_admin(c.from_user.id, c.from_user.username):
        await safe_answer(c, "Только для админов.", show_alert=True); return
    async with Session() as session:
        res = await session.execute(select(Player).order_by(Player.first_name.asc(), Player.last_name.asc()))
        all_players = list(res.scalars().all())
    ids = _load_day_list()
    await safe_edit(c.message, "Настройка «Списка дня». Отметьте игроков и нажмите «Сохранить список».", reply_markup=daylist_kb(all_players, ids))
    await safe_answer(c, )

@dp.callback_query(F.data.startswith("day:toggle:"))
async def day_toggle(c: CallbackQuery):
    metric_click(c.from_user.id)
    if not is_admin(c.from_user.id, c.from_user.username):
        await safe_answer(c, "Только для админов.", show_alert=True); return
    try:
        pid = int(c.data.split(":")[2])
    except Exception:
        await safe_answer(c, ); return
    ids = _load_day_list()
    if pid in ids:
        ids = [i for i in ids if i != pid]
    else:
        ids = ids + [pid]
    # сверим с БД и сохраним
    async with Session() as session:
        res = await session.execute(select(Player.id))
        valid_ids = set(res.scalars().all())
    ids = [i for i in ids if i in valid_ids]
    _save_day_list(ids)
    # перерисуем клавиатуру
    async with Session() as session:
        res = await session.execute(select(Player).order_by(Player.first_name.asc(), Player.last_name.asc()))
        all_players = list(res.scalars().all())
    await safe_edit(c.message, "Настройка «Списка дня». Отметьте игроков и нажмите «Сохранить список».", reply_markup=daylist_kb(all_players, ids))
    await safe_answer(c, )

@dp.callback_query(F.data == "day:clear")
async def day_clear(c: CallbackQuery):
    metric_click(c.from_user.id)
    if not is_admin(c.from_user.id, c.from_user.username):
        await safe_answer(c, "Только для админов.", show_alert=True); return
    _save_day_list([])
    async with Session() as session:
        res = await session.execute(select(Player).order_by(Player.first_name.asc(), Player.last_name.asc()))
        all_players = list(res.scalars().all())
    await safe_edit(c.message, "Список дня очищен.", reply_markup=daylist_kb(all_players, []))
    await safe_answer(c, "Очищено.")

@dp.callback_query(F.data == "day:save")
async def day_save(c: CallbackQuery):
    metric_click(c.from_user.id)
    if not is_admin(c.from_user.id, c.from_user.username):
        await safe_answer(c, "Только для админов.", show_alert=True); return
    # ничего не делаем: список уже сохранён на каждом клике, просто сообщаем
    await safe_answer(c, "Сохранено.")

# ===================== start / faq =====================
FAQ_TEXT = (
"""❓ *FAQ*

Просьба не путать очки общего рейтинга *(MMR — изначально у каждого 3000 MMR)* и *социальные очки* для выявления лучшего синего/красного/Воланда *(изначально у всех 0 очков)*.

1) *Как бот определяет рейтинги команд и кто сильнее?* — по *среднему MMR* игроков в команде.
2) *Если команда A намного сильнее команды B?* — при большой разнице сильная команда при победе получает меньше MMR, при поражении теряет больше.
3) *Социальные очки* — начисляются только победившей стороне согласно исходу матча.

*MMR — простая формула дельты:*
1) diff = разница средних MMR команд.
2) x = floor(diff/10), если diff > 400 → x = 41.
3) Если победила сильная: + (51−x) сильной и − (49−x) слабой.
4) Если победила слабая: + (51+x) слабой и − (49+x) сильной.
Потолки при большой разнице: +10/−8 (победа сильной) и +92/−90 (победа слабой).

*Галлеоны (внутриигровая валюта)*
1) Система начисления Галлеонов 🪙. Игроку зачисляется:
- 1 монета за участие в любой игре (независимо от команды)
- 1 монета за победу в игре 
- Если игрок избран Воландемортом он получает ещё 3 монеты сверху (в итоге 5: 1 за участие + 1 за победу + 3 избрание)
- Если игрок убивает Воландеморта — ещё 5 монет сверху (в итоге 1 за участие + 1 за победу + 5 за убийство)

"Винстрик"
- Определение: <i>винстрик — это количество побед подряд</i>. После первой победы лузстрик сбрасывается в 0, а активный винстрик становится 1.
- Если игрок побеждает 2 раза подряд — ему зачисляется на баланс 2 монеты (просто добавляется 2 монеты к его балансу)
- Если игрок побеждает 3 раза подряд — ему зачисляется на баланс 4 монеты
- Если игрок побеждает 4 раза подряд — ему зачисляется на баланс 8 монет
- Если игрок побеждает 5 раз подряд — ему зачисляется на баланс 16 монет
- Если игрок побеждает 6 раз подряд — ему зачисляется на баланс 32 монеты
- Если игрок побеждает 7 раз подряд — ему зачисляется на баланс 100 монет
Если игрок продолжает побеждать без поражений, ему зачисляется по 100 монет сверху за каждую победу (плюс стандартные +1 за участие и +1 за победу).

"Лузстрик"
- Определение: <i>лузстрик — это количество поражений подряд</i>. После первого поражения винстрик сбрасывается в 0, а активный лузстрик становится 1.
- Если игрок проигрывает 2 раза подряд — ему зачисляется на баланс 2 монеты
- Если игрок проигрывает 4 раза подряд — ему зачисляется на баланс 4 монеты
- Если игрок проигрывает 6 раз подряд — ему зачисляется на баланс 6 монет
Если игрок проигрывает 6 и более раз подряд, за каждое следующее поражение ему даётся по 6 монет сверху (+1 за участие), пока не будет хотя бы 1 победа; далее «стрик поражений» сбрасывается."""
)
@dp.message(CommandStart())
async def start_cmd(m: Message, state: FSMContext):
    await state.clear()
    metric_visit(m.from_user.id)
    admin = is_admin(m.from_user.id, m.from_user.username)
    authorized = is_authorized_user(m.from_user.id)
    await m.answer(
        "Главное меню.\nЭтот бот ведёт рейтинги игры «Тайный Воландеморт».",
        reply_markup=home_kb_for_user(admin, authorized),
    )

@dp.callback_query(F.data == "faq")
async def faq(c: CallbackQuery, state: FSMContext):
    metric_click(c.from_user.id)
    if await _maybe_warn_unfinished(c, state, "faq"):
        return
    admin = is_admin(c.from_user.id, c.from_user.username)
    await safe_edit(
        c.message,
        FAQ_TEXT,
        parse_mode="Markdown",
        reply_markup=home_kb_for_user(admin, is_authorized_user(c.from_user.id)),
    )
    await safe_answer(c, )

@dp.callback_query(F.data == "backhome")
async def back_home(c: CallbackQuery, state: FSMContext):
    metric_click(c.from_user.id)
    if await _maybe_warn_unfinished(c, state, "backhome"):
        return
    await state.clear()
    admin = is_admin(c.from_user.id, c.from_user.username)
    await safe_edit(
        c.message,
        "Главное меню.\nЭтот бот ведёт рейтинги игры «Тайный Воландеморт».",
        reply_markup=home_kb_for_user(admin, is_authorized_user(c.from_user.id)),
    )
    await safe_answer(c, )

# ===================== Authorization =====================
@dp.callback_query(F.data == "auth:start")
async def auth_start(c: CallbackQuery, state: FSMContext):
    metric_click(c.from_user.id)
    if await _maybe_warn_unfinished(c, state, "auth:start"):
        return
    await state.set_state(UserAuthFSM.wait_name)
    admin = is_admin(c.from_user.id, c.from_user.username)
    admin_note = "\n\n_Вы админ — заявка будет одобрена автоматически._" if admin else ""
    await safe_edit(
        c.message,
        "Привет! Чтобы получить полный доступ к боту и бороться за рейтинг, *напиши своё имя и фамилию* ❤️\n\n"
        "_Одним сообщением, до 25 символов._" + admin_note,
        parse_mode="Markdown",
    )
    await safe_answer(c, )

@dp.message(UserAuthFSM.wait_name)
async def auth_take_name(m: Message, state: FSMContext):
    raw = (m.text or "").strip()
    if not raw or len(raw) > 25:
        await m.answer("Пожалуйста, отправьте одно сообщение с именем и фамилией (до 25 символов).")
        return

    if is_authorized_user(m.from_user.id):
        await state.clear()
        await m.answer("Вы уже авторизованы. Откройте главное меню или «📊 Моя статистика».")
        return

    if is_admin(m.from_user.id, m.from_user.username):
        parts = raw.split()
        first, last = parts[0], (" ".join(parts[1:]) if len(parts) > 1 else None)
        async with Session() as session:
            p = await create_player(session, first_name=first, last_name=last, username=m.from_user.username or None)
            link_user_to_player(m.from_user.id, p.id)
        await state.clear()
        await m.answer("Готово! Вам открыт полный доступ к боту. Удачных игр ❤️",
                       reply_markup=home_kb_for_user(True, True))
        return

    apps = _load_apps()
    apps = [a for a in apps if not (a.get("user_id")==m.from_user.id and a.get("status")=="pending")]
    apps.append({
        "user_id": m.from_user.id,
        "chat_id": m.chat.id,
        "tg_username": m.from_user.username or None,
        "name": raw,
        "status": "pending",
    })
    _save_apps(apps)
    await state.clear()
    await m.answer("Заявка отправлена администратору. Ожидайте подтверждения 🙌")

# ===================== Game creation (skip title, show MSK time) =====================
@dp.callback_query(F.data == "start:newgame")
async def start_newgame(c: CallbackQuery, state: FSMContext):
    if not is_admin(c.from_user.id, c.from_user.username):
        await safe_answer(c, "Только для админов.", show_alert=True); return
    metric_click(c.from_user.id)
    async with Session() as session:
        ts = now_msk()
        # В заголовок кладём человекочитаемое время по МСК
        title = f"Игра от {ts.strftime('%d.%m.%Y %H:%M')} (МСК)"
        g = await create_game(session, title, user_id=c.from_user.id)
        metric_inc("games_created")
    await state.update_data(pending_gid=g.id)
    await state.update_data(pending_gid=g.id)
    await state.set_state(CreateGameFSM.main_menu)
    await safe_edit(
        c.message,
        f"Создана: *{title}* (ID {g.id}). Заполните команды и Воландеморта.",
        parse_mode="Markdown",
        reply_markup=main_menu_kb(g.id),
    )
    await safe_answer(c, )

# ===================== Back to menu of a game =====================
@dp.callback_query(F.data.startswith("back:"))
async def back_to_menu(c: CallbackQuery, state: FSMContext):
    metric_click(c.from_user.id)
    game_id = int(c.data.split(":")[1])
    async with Session() as session:
        summary, *_ = await roster_summary(session, game_id)
        g = await get_game(session, game_id)
    await safe_edit(
        c.message,
        f"Игра: *{getattr(g,'title','Игра')}*.\n\n{summary}",
        parse_mode="Markdown",
        reply_markup=main_menu_kb(game_id),
    )
    await safe_answer(c, )

# ===================== Pick teams / selection =====================
async def effective_limit(session: Session, team: str, game_id: int) -> int:
    return MAX_BLUE if team == "blue" else 3

@dp.callback_query(F.data.startswith("multiteam:"))
async def multiteam_entry(c: CallbackQuery, state: FSMContext):
    metric_click(c.from_user.id)
    _, team, game_id_s = c.data.split(":")
    game_id = int(game_id_s)
    if is_admin(c.from_user.id, c.from_user.username):
        await safe_edit(c.message, "Выберите источник списка игроков:", reply_markup=source_choice_kb(team, game_id))
        await safe_answer(c, )
        return
    await _open_multiteam_with_source(c, state, team, game_id, source="all")

@dp.callback_query(F.data.startswith("source:"))
async def choose_source(c: CallbackQuery, state: FSMContext):
    metric_click(c.from_user.id)
    _, src, team, game_id_s = c.data.split(":")
    if team == "voldemort":
        await _open_vold_with_source(c, int(game_id_s), source=src)
    else:
        await _open_multiteam_with_source(c, state, team, int(game_id_s), source=src)

async def _open_multiteam_with_source(c: CallbackQuery, state: FSMContext, team: str, game_id: int, source: str):
    async with Session() as session:
        if source == "day":
            ids = _load_day_list()
            if not ids:
                await safe_answer(c, "«Список дня» пуст. Отметьте игроков в админ-панели.", show_alert=True)
                source = "all"
        if source == "day":
            res = await session.execute(
                select(Player).where(Player.id.in_(ids)).order_by(Player.first_name.asc(), Player.last_name.asc())
            )
        else:
            res = await session.execute(
                select(Player).order_by(Player.first_name.asc(), Player.last_name.asc())
            )
        players = list(res.scalars().all())
        blue, red, vold = await get_team_rosters(session, game_id)
        g = await get_game(session, game_id)
        vold_id = g.voldemort_id if g else None
        selected_ids = [p.id for p in (blue if team == "blue" else [p for p in red if not (vold and p.id == vold.id)])]
        limit = await effective_limit(session, team, game_id)
        blue_ids = [p.id for p in blue]
        red_ids = [p.id for p in red if not (vold and p.id == vold.id)]

    await state.update_data(game_id=game_id, select_team=team, selected_ids=selected_ids)
    header = f"{'🔵' if team == 'blue' else '🔴'} Выбор игроков ({'список дня' if source=='day' else 'все'}) — {len(selected_ids)}/{limit}"
    await safe_edit(
        c.message,
        header,
        reply_markup=multiselect_kb(
            players, selected_ids, team, game_id, limit, vold_id,
            admin_can_add=False,  # ⛔️ режем создание игрока из набора
            blue_ids=blue_ids, red_ids=red_ids
        )
    )
    await state.update_data(source=source)
    await state.set_state(CreateGameFSM.selecting_team)
    await safe_answer(c, )

@dp.callback_query(F.data.startswith("toggle:"))
async def toggle_player(c: CallbackQuery, state: FSMContext):
    metric_click(c.from_user.id)
    _, team, game_id_s, player_id_s = c.data.split(":")
    game_id = int(game_id_s)
    pid = int(player_id_s)
    data = await state.get_data()
    selected_ids: List[int] = data.get("selected_ids", [])

    async with Session() as session:
        blue, red, vold = await get_team_rosters(session, game_id)
        g = await get_game(session, game_id)
        vold_id = g.voldemort_id if g else None
        blue_ids = [p.id for p in blue]
        red_ids = [p.id for p in red if not (vold and p.id == vold.id)]
        if team == "blue":
            if pid in red_ids or (vold_id and pid == vold_id):
                await safe_answer(c, "Этот игрок уже в красных/он Воландеморт.", show_alert=True)
                return
        else:
            if pid in blue_ids or (vold_id and pid == vold_id):
                await safe_answer(c, "Этот игрок уже в синих или является Воландемортом.", show_alert=True)
                return

    if pid in selected_ids:
        selected_ids = [x for x in selected_ids if x != pid]
    else:
        async with Session() as session2:
            limit = await effective_limit(session2, team, game_id)
        if len(selected_ids) >= limit:
            await safe_answer(c, f"Достигнут лимит: {limit}.", show_alert=True)
            return
        selected_ids.append(pid)

    await state.update_data(selected_ids=selected_ids)

    async with Session() as session:
        src = (await state.get_data()).get("source", "all")
        if src == "day":
            ids = _load_day_list()
            res_ids = await session.execute(select(Player.id))
            valid_ids = set(res_ids.scalars().all())
            ids = [i for i in ids if i in valid_ids]
            if ids:
                res = await session.execute(select(Player).where(Player.id.in_(ids)).order_by(Player.first_name.asc(), Player.last_name.asc()))
            else:
                res = await session.execute(select(Player).order_by(Player.first_name.asc(), Player.last_name.asc()).limit(500))
        else:
            res = await session.execute(select(Player).order_by(Player.first_name.asc(), Player.last_name.asc()).limit(500))
        players = list(res.scalars().all())
        g = await get_game(session, game_id)
        vold_id = g.voldemort_id if g else None
        limit = await effective_limit(session, team, game_id)
        blue, red, vold = await get_team_rosters(session, game_id)
        blue_ids = [p.id for p in blue]
        red_ids = [p.id for p in red if not (vold and p.id == vold.id)]
        header = f"{'🔵' if team == 'blue' else '🔴'} Выбрано: {len(selected_ids)} / {limit}"
        await safe_edit(
            c.message,
            header,
            reply_markup=multiselect_kb(
                players, selected_ids, team, game_id, limit, vold_id,
                admin_can_add=False,
                blue_ids=blue_ids, red_ids=red_ids
            )
        )
        await safe_answer(c, )

@dp.callback_query(F.data.startswith("clear:"))
async def clear_selection(c: CallbackQuery, state: FSMContext):
    metric_click(c.from_user.id)
    _, team, game_id_s = c.data.split(":")
    game_id = int(game_id_s)
    await state.update_data(selected_ids=[])
    async with Session() as session:
        src = (await state.get_data()).get("source", "all")
        if src == "day":
            ids = _load_day_list()
            res_ids = await session.execute(select(Player.id))
            valid_ids = set(res_ids.scalars().all())
            ids = [i for i in ids if i in valid_ids]
            if ids:
                res = await session.execute(select(Player).where(Player.id.in_(ids)).order_by(Player.first_name.asc(), Player.last_name.asc()))
            else:
                res = await session.execute(select(Player).order_by(Player.first_name.asc(), Player.last_name.asc()).limit(500))
        else:
            res = await session.execute(select(Player).order_by(Player.first_name.asc(), Player.last_name.asc()).limit(500))
        players = list(res.scalars().all())
        limit = await effective_limit(session, team, game_id)
        g = await get_game(session, game_id)
        vold_id = g.voldemort_id if g else None
        blue, red, vold = await get_team_rosters(session, game_id)
        blue_ids = [p.id for p in blue]
        red_ids = [p.id for p in red if not (vold and p.id == vold.id)]
    header = f"{'🔵' if team == 'blue' else '🔴'} Выбрано: 0 / {limit}"
    await safe_edit(
        c.message,
        header,
        reply_markup=multiselect_kb(
            players, [], team, game_id, limit, vold_id,
            admin_can_add=False,
            blue_ids=blue_ids, red_ids=red_ids
        )
    )
    await safe_answer(c, "Сброшено.")

@dp.callback_query(F.data.startswith("save:"))
async def save_selection(c: CallbackQuery, state: FSMContext):
    metric_click(c.from_user.id)
    _, team, game_id_s = c.data.split(":")
    game_id = int(game_id_s)
    data = await state.get_data()
    selected_ids: List[int] = data.get("selected_ids", [])
    async with Session() as session:
        await set_team_roster(session, game_id, team, selected_ids)
        summary, _, _, _ = await roster_summary(session, game_id)
    await safe_answer(c, "Команда сохранена.")
    await safe_edit(
        c.message,
        f"Состав сохранён.\n\n{summary}",
        reply_markup=main_menu_kb(game_id),
        parse_mode="Markdown",
    )

# ===================== Voldemort =====================
@dp.callback_query(F.data.startswith("vold:"))
async def choose_voldemort_entry(c: CallbackQuery, state: FSMContext):
    metric_click(c.from_user.id)
    _, game_id_s = c.data.split(":")
    game_id = int(game_id_s)
    if is_admin(c.from_user.id, c.from_user.username):
        await safe_edit(c.message, "Выберите источник списка игроков (Воландеморт):", reply_markup=source_choice_kb("voldemort", game_id))
        await safe_answer(c, )
        return
    await _open_vold_with_source(c, game_id, source="all")

async def _open_vold_with_source(c: CallbackQuery, game_id: int, source: str):
    async with Session() as session:
        if source == "day":
            ids = _load_day_list()
            if not ids:
                await safe_answer(c, "«Список дня» пуст. Отметьте игроков в админ-панели.", show_alert=True)
                source = "all"
        if source == "day":
            res = await session.execute(
                select(Player).where(Player.id.in_(_load_day_list())).order_by(Player.first_name.asc(), Player.last_name.asc())
            )
        else:
            res = await session.execute(
                select(Player).order_by(Player.first_name.asc(), Player.last_name.asc())
            )
        players = list(res.scalars().all())
        blue, red, vold = await get_team_rosters(session, game_id)
        blue_ids = [p.id for p in blue]
        red_ids = [p.id for p in red if not (vold and p.id == vold.id)]
        vold_id = vold.id if vold else None
    kb = InlineKeyboardBuilder()
    for p in players:
        prefix = _status_prefix(p.id, [], blue_ids, red_ids, vold_id, "🟣")
        kb.button(text=f"{prefix}{full_name(p)} [{p.rating}]", callback_data=f"pickv:{game_id}:{p.id}")
    kb.button(text="🔎 Поиск", callback_data=f"search:voldemort:{game_id}")
    kb.button(text="⬅️ Назад", callback_data=f"back:{game_id}")
    kb.adjust(1)
    await safe_edit(c.message, "Выберите Воландеморта (🟣). Он не должен быть в синих.", reply_markup=kb.as_markup())

@dp.callback_query(F.data.startswith("pickv:"))
async def pick_voldemort(c: CallbackQuery):
    metric_click(c.from_user.id)
    _, game_id_s, player_id_s = c.data.split(":")
    game_id = int(game_id_s)
    pid = int(player_id_s)
    async with Session() as session:
        blue, red, vold = await get_team_rosters(session, game_id)
        if pid in [p.id for p in blue]:
            await safe_answer(c, "Этот игрок уже в синих — уберите его из синих сначала.", show_alert=True)
            return
        await set_voldemort(session, game_id, pid)
        summary, *_ = await roster_summary(session, game_id)
        g = await get_game(session, game_id)
    await safe_edit(
        c.message,
        f"Игра: *{getattr(g,'title','Игра')}*.\n\n{summary}",
        parse_mode="Markdown",
        reply_markup=main_menu_kb(game_id),
    )
    await safe_answer(c, "Воландеморт задан.")

# ===================== Search (без «создать игрока») =====================
@dp.callback_query(F.data.startswith("search:"))
async def ask_search(c: CallbackQuery, state: FSMContext):
    metric_click(c.from_user.id)
    _, team, game_id_s = c.data.split(":")
    await state.update_data(search_target=team, game_id=int(game_id_s), _return_to="teamselect")
    await safe_edit(c.message, "Введите имя или имя+фамилию (например, *Иван Петров*):", parse_mode="Markdown")
    await state.set_state(CreateGameFSM.search_player_for)
    await safe_answer(c, )

@dp.message(CreateGameFSM.search_player_for)
async def search_players_msg(m: Message, state: FSMContext):
    data = await state.get_data()
    team = data.get("search_target")
    game_id = data.get("game_id")
    query = (m.text or "").strip()
    async with Session() as session:
        players = await search_players(session, query)
        g = await get_game(session, game_id)
        vold_id = g.voldemort_id if g else None
        blue, red, vold = await get_team_rosters(session, game_id)
        blue_ids = [p.id for p in blue]
        red_ids = [p.id for p in red if not (vold and p.id == vold.id)]
    if not players:
        kb = InlineKeyboardBuilder()
        kb.button(text="Попробовать ещё", callback_data=f"search:{team}:{game_id}")
        kb.button(text="⬅️ Назад", callback_data=f"back:{game_id}")
        kb.adjust(1)
        await m.answer("Ничего не найдено.", reply_markup=kb.as_markup())
        return
    if team in ("blue", "red"):
        selected_ids = data.get("selected_ids", [])
        async with Session() as session2:
            limit = await effective_limit(session2, team, game_id)
        header = f"{'🔵' if team == 'blue' else '🔴'} Выбрано: {len(selected_ids)} / {limit}"
        await m.answer(
            header,
            reply_markup=multiselect_kb(
                players, selected_ids, team, game_id, limit, vold_id,
                admin_can_add=False, blue_ids=blue_ids, red_ids=red_ids
            ),
        )
    else:
        kb = InlineKeyboardBuilder()
        for p in players:
            prefix = _status_prefix(p.id, [], blue_ids, red_ids, vold_id, "🟣")
            kb.button(text=f"{prefix}{full_name(p)} [{p.rating}]", callback_data=f"pickv:{game_id}:{p.id}")
        kb.button(text="⬅️ Назад", callback_data=f"back:{game_id}")
        kb.adjust(1)
        await m.answer("Результаты поиска (Воландеморт):", reply_markup=kb.as_markup())

# ===================== Check / Winner / Apply ratings =====================
@dp.callback_query(F.data.startswith("check:"))
async def check_roster(c: CallbackQuery, state: FSMContext):
    metric_click(c.from_user.id)
    _, game_id_s = c.data.split(":")
    game_id = int(game_id_s)
    async with Session() as session:
        summary, *_ = await roster_summary(session, game_id)
    await safe_edit(c.message, summary, reply_markup=main_menu_kb(game_id))
    await safe_answer(c, )

@dp.callback_query(F.data.startswith("winner:"))
async def choose_winner(c: CallbackQuery, state: FSMContext):
    metric_click(c.from_user.id)
    _, game_id_s = c.data.split(":")
    game_id = int(game_id_s)
    async with Session() as session:
        blue, red, vold = await get_team_rosters(session, game_id)
        ok, msg = await validate_rosters(blue, red, vold)
    if not ok:
        await safe_answer(c, msg, show_alert=True)
        return
    kb = InlineKeyboardBuilder()
    kb.button(text="🟦 Победа Ордена Феникса — 5 законов", callback_data=f"setres:blue_laws:{game_id}")
    kb.button(text="🟦 Воландеморт убит (Орден Феникса)", callback_data=f"setres:blue_kill:{game_id}")
    kb.button(text="🟥 Победа Пожирателей — 6 законов", callback_data=f"setres:red_laws:{game_id}")
    kb.button(text="🟥 Воландеморт директор (Пожиратели)", callback_data=f"setres:red_director:{game_id}")
    kb.button(text="⬅️ Назад", callback_data=f"back:{game_id}")
    kb.adjust(1)
    await safe_edit(c.message, "Выберите исход игры:", reply_markup=kb.as_markup())
    await safe_answer(c, )

@dp.callback_query(F.data.startswith("setres:"))
async def set_result(c: CallbackQuery, state: FSMContext):
    metric_click(c.from_user.id)
    _, result_type, game_id_s = c.data.split(":")
    game_id = int(game_id_s)

    if result_type == "blue_kill":
        async with Session() as session:
            blue, red, vold = await get_team_rosters(session, game_id)
            if not vold:
                await safe_answer(c, "Сначала выберите Воландеморта.", show_alert=True)
                return
            res = await session.execute(
                select(Player).where(Player.id.in_([p.id for p in blue])).order_by(Player.first_name.asc(), Player.last_name.asc())
            )
            blue_sorted = list(res.scalars().all())
        kb = InlineKeyboardBuilder()
        for p in blue_sorted:
            kb.button(text=f"🗡️ {full_name(p)}", callback_data=f"killpick:{game_id}:{p.id}")
        kb.button(text="⬅️ Назад", callback_data=f"winner:{game_id}")
        kb.adjust(1)
        await state.update_data(pending_result=result_type, game_id=game_id)
        await safe_edit(c.message, "Кто убил Воландеморта? Выберите игрока:", reply_markup=kb.as_markup())
        await state.set_state(CreateGameFSM.wait_pick_killer)
        await safe_answer(c, )
        return

    async with Session() as session:
        await set_result_type_and_killer(session, game_id, result_type, killer_id=None)
        await state.update_data(pending_gid=None)
        await state.update_data(pending_gid=None)
        summary = await apply_ratings(session, game_id)
        summary = _normalize_summary_delta(summary)
        summary = _normalize_summary_delta(summary)
        summary = _strip_repeat_summary(summary)
        summary = _strip_repeat_summary(summary)
        blue, red, vold = await get_team_rosters(session, game_id)
        # include Voldemort into red side for averages
        red_ext = list(red)
        if vold and all(p.id != vold.id for p in red_ext):
            red_ext.append(vold)
        b_avg = round(sum(p.rating for p in blue) / max(1, len(blue)), 1)
        red_ext = list(red)
        if vold and all(p.id != vold.id for p in red_ext):
            red_ext.append(vold)
        r_avg = round(sum(p.rating for p in red_ext) / max(1, len(red_ext)), 1)
        fav = favorite_side(b_avg, r_avg)
        metric_inc("games_finished")

    human = RESULT_HUMAN.get(result_type, "Исход не указан")
    await safe_edit(
        c.message,
        f"Игра завершена.\n"
        f"{human}\n"
        f"Средний MMR — Орден Феникса: {b_avg}, Пожиратели: {r_avg}\n"
        f"Фаворит матча: {fav}\n"
        f"{summary}",
        reply_markup=after_finish_kb(),
    )
    await safe_answer(c, )

@dp.callback_query(F.data.startswith("killpick:"))
async def picked_killer(c: CallbackQuery, state: FSMContext):
    metric_click(c.from_user.id)
    _, game_id_s, killer_id_s = c.data.split(":")
    game_id = int(game_id_s)
    killer_id = int(killer_id_s)
    async with Session() as session:
        await set_result_type_and_killer(session, game_id, "blue_kill", killer_id=killer_id)
        summary = await apply_ratings(session, game_id)
        summary = _normalize_summary_delta(summary)
        summary = _normalize_summary_delta(summary)
        blue, red, vold = await get_team_rosters(session, game_id)
        # include Voldemort into red side for averages
        red_ext = list(red)
        if vold and all(p.id != vold.id for p in red_ext):
            red_ext.append(vold)
        b_avg = round(sum(p.rating for p in blue) / max(1, len(blue)), 1)
        r_avg = round(sum(p.rating for p in red_ext) / max(1, len(red_ext)), 1)
        fav = favorite_side(b_avg, r_avg)
        metric_inc("games_finished")

    await state.clear()
    await safe_edit(
        c.message,
        "Игра завершена.\n"
        f"{RESULT_HUMAN['blue_kill']}\n"
        f"Средний MMR — Орден Феникса: {b_avg}, Пожиратели: {r_avg}\n"
        f"Фаворит матча: {fav}\n"
        f"{summary}",
        reply_markup=after_finish_kb(),
    )
    await safe_answer(c, "Киллер сохранён.")

# ===================== Ratings / Export / Tops =====================
@dp.callback_query(F.data == "rating:menu")
async def rating_menu(c: CallbackQuery, state: FSMContext):
    metric_click(c.from_user.id)
    if await _maybe_warn_unfinished(c, state, "rating:menu"):
        return
    async with Session() as session:
        res = await session.execute(
            select(Player).order_by(Player.rating.desc(), Player.first_name.asc(), Player.last_name.asc()).limit(100)
        )
        players = list(res.scalars().all())
    if not players:
        admin = is_admin(c.from_user.id, c.from_user.username)
        await safe_edit(c.message, "Пока нет игроков.", reply_markup=home_kb_for_user(admin, is_authorized_user(c.from_user.id)))
        await safe_answer(c, )
        return
    lines = [f"{i+1}. {full_name(p)} — {p.rating}" for i, p in enumerate(players)]
    await safe_edit(
        c.message,
        "🏆 *Рейтинг игроков (топ 100)*\n\n" + "\n".join(lines),
        parse_mode="Markdown",
        reply_markup=rating_kb(),
    )
    await safe_answer(c, )

@dp.callback_query(F.data == "rating:export")
async def rating_export(c: CallbackQuery, state: FSMContext):
    metric_click(c.from_user.id)
    metric_inc("excel_downloads")
    from openpyxl import Workbook

    async with Session() as session:
        from services import recompute_win_counters
        await recompute_win_counters(session)
        res = await session.execute(
            select(Player).order_by(Player.rating.desc(), Player.first_name.asc(), Player.last_name.asc())
        )
        players = list(res.scalars().all())

    with tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx") as tmp:
        file_path = tmp.name
    try:
        wb = Workbook()
        ws = wb.active
        ws.title = "Рейтинг"
        ws.append(["#", "Имя", "Фамилия", "MMR", "Победы Ордена", "Победы Пожирателей (вкл. Воландеморта)", "Директором избран Воландеморт", "Игрок отправил Воландеморта в Азкабан"])
        for i, p in enumerate(players, start=1):
            ws.append([i, p.first_name, (p.last_name or ""), int(p.rating), int(getattr(p, "blue_wins", 0) or 0), int(getattr(p, "red_wins", 0) or 0) + int(getattr(p, "vold_wins", 0) or 0), int(getattr(p, "social_vold", 0) or 0), int(getattr(p, "killer_points", 0) or 0)])
        wb.save(file_path)
        await c.message.answer_document(FSInputFile(file_path), caption="Экспорт рейтинга (Excel)")
        await safe_answer(c, "Файл готов.")
    finally:
        try: os.remove(file_path)
        except Exception: pass

@dp.callback_query(F.data.startswith("rating:top:"))
async def rating_top(c: CallbackQuery):
    metric_click(c.from_user.id)
    _, _, role = c.data.split(":")
    titles = {
        "blue": "🌟 Лучшие синие (по социальным очкам)",
        "red": "🔥 Лучшие пожиратели (по социальным очкам)",
        "vold": "🕶️ Лучшие Воландеморты (по социальным очкам)",
        "killer": "🗡️ Лучшие киллеры Воландеморта (по убийствам)",
    }
    async with Session() as session:
        if role == "blue":
            res = await session.execute(
                select(Player).order_by(desc(Player.social_blue), Player.first_name, Player.last_name).limit(50)
            )
        elif role == "red":
            res = await session.execute(
                select(Player).order_by(desc(Player.social_red), Player.first_name, Player.last_name).limit(50)
            )
        elif role == "vold":
            res = await session.execute(
                select(Player).order_by(desc(Player.social_vold), Player.first_name, Player.last_name).limit(50)
            )
        else:
            res = await session.execute(
                select(Player).order_by(desc(Player.killer_points), Player.first_name, Player.last_name).limit(50)
            )
        players = list(res.scalars().all())

    if not players:
        await safe_edit(c.message, "Пока нет статистики.", reply_markup=rating_kb()); await safe_answer(c, ); return

    def points(p: Player) -> int:
        return (p.social_blue if role=="blue"
                else p.social_red if role=="red"
                else p.social_vold if role=="vold"
                else p.killer_points)

    lines = [f"{i+1}. {full_name(p)} — {points(p)} очк." for i, p in enumerate(players)]
    await safe_edit(c.message, f"{titles[role]}\n\n" + "\n".join(lines), reply_markup=rating_kb())
    await safe_answer(c, )

# ===================== Finished games (admin & users) =====================

def botstats_menu_kb():
    kb = InlineKeyboardBuilder()
    kb.button(text="За неделю", callback_data="botstats:week")
    kb.button(text="За месяц", callback_data="botstats:month")
    kb.button(text="За всё время", callback_data="botstats:all")
    kb.button(text="⬅️ Назад", callback_data="admin:menu")
    kb.adjust(1)
    return kb.as_markup()
def finished_menu_kb():
    kb = InlineKeyboardBuilder()
    kb.button(text="🗓 За последнюю неделю", callback_data="finished:week")
    kb.button(text="📚 Все игры", callback_data="finished:all")
    kb.button(text="⬅️ В главное меню", callback_data="backhome")
    kb.adjust(1)
    return kb.as_markup()

def games_pick_kb(items: List[Game], allow_notes: bool):
    kb = InlineKeyboardBuilder()
    for g in items:
        feather = " 🖋️" if _has_notes(g.id) else ""
        title = getattr(g, "title", f"Игра {g.id}")  # в title уже включено время по МСК
        kb.button(text=f"ID {g.id}: {title}{feather}", callback_data=f"finished:view:{g.id}")
    kb.button(text="⬅️ Назад", callback_data="finished:menu")
    kb.adjust(1)
    return kb.as_markup()

def finished_actions_kb(game_id: int, admin: bool):
    kb = InlineKeyboardBuilder()
    kb.button(text="👀 Посмотреть результаты", callback_data=f"finished:result:{game_id}")
    if admin:
        kb.button(text="🖊 Оставить заметку", callback_data=f"finished:note:{game_id}")
    kb.button(text="⬅️ Назад", callback_data="finished:menu")
    kb.adjust(1)
    return kb.as_markup()


# ---- Leave confirmation handlers ----
@dp.callback_query(F.data.startswith("leave:confirm:"))
async def leave_confirm(c: CallbackQuery, state: FSMContext):
    _, _, gid_s, target_enc = c.data.split(":", 3)
    gid = int(gid_s)
    target = _decode_target(target_enc)
    async with Session() as session:
        try:
            await delete_game(session, gid)
        except Exception:
            pass
    await state.clear()
    # route
    if target == "finished:menu":
        return await finished_menu(c, state)  # type: ignore
    if target == "rating:menu":
        return await rating_menu(c, state)  # type: ignore
    if target == "playeroftheday":
        return await player_of_the_day(c, state)  # type: ignore
    if target == "faq":
        return await faq(c, state)  # type: ignore
    if target == "auth:start":
        return await auth_start(c, state)  # type: ignore
    if target == "admin:menu":
        return await admin_menu(c, state)  # type: ignore
    if target == "backhome":
        return await back_home(c, state)  # type: ignore
    if target == "me:stats":
        return await my_stats(c, state)  # type: ignore
    return await back_home(c, state)

@dp.callback_query(F.data.startswith("leave:stay:"))
async def leave_stay(c: CallbackQuery, state: FSMContext):
    _, _, gid_s = c.data.split(":")
    gid = int(gid_s)
    async with Session() as session:
        summary, *_ = await roster_summary(session, gid)
        g = await get_game(session, gid)
    await safe_edit(
        c.message,
        f"Игра: *{getattr(g,'title','Игра')}*\n\n{summary}",
        parse_mode="Markdown",
        reply_markup=main_menu_kb(gid),
    )
    await safe_answer(c, )
@dp.callback_query(F.data == "finished:menu")
async def finished_menu(c: CallbackQuery, state: FSMContext):
    metric_click(c.from_user.id)
    if await _maybe_warn_unfinished(c, state, "finished:menu"):
        return
    await safe_edit(c.message, "Завершённые игры — выберите период:", reply_markup=finished_menu_kb())
    await safe_answer(c, )

def _games_in_range(all_games: List[Game], start: Optional[datetime]) -> List[Game]:
    if start is None:
        return all_games
    start_date = start.date()
    out = []
    for g in all_games:
        created = getattr(g, "created_at", None)
        if not created:
            out.append(g); continue
        try:
            created_date = created.date()  # сравниваем только даты
        except Exception:
            out.append(g); continue
        if created_date >= start_date:
            out.append(g)
    return out

@dp.callback_query(F.data == "finished:week")
async def finished_week(c: CallbackQuery):
    metric_click(c.from_user.id)
    async with Session() as session:
        games = await list_all_games(session)
    week_ago = now_msk() - timedelta(days=7)
    items = _games_in_range(games, week_ago)
    if not items:
        await safe_edit(c.message, "За последнюю неделю игр нет.", reply_markup=finished_menu_kb()); await safe_answer(c, ); return
    await safe_edit(c.message, "Выберите игру:", reply_markup=games_pick_kb(items, allow_notes=is_admin(c.from_user.id, c.from_user.username)))
    await safe_answer(c, )

@dp.callback_query(F.data == "finished:all")
async def finished_all(c: CallbackQuery):
    metric_click(c.from_user.id)
    async with Session() as session:
        games = await list_all_games(session)
    if not games:
        await safe_edit(c.message, "Игр ещё нет.", reply_markup=finished_menu_kb()); await safe_answer(c, ); return
    await safe_edit(c.message, "Выберите игру:", reply_markup=games_pick_kb(games, allow_notes=is_admin(c.from_user.id, c.from_user.username)))
    await safe_answer(c, )

@dp.callback_query(F.data.startswith("finished:view:"))
async def finished_view(c: CallbackQuery):
    metric_click(c.from_user.id)
    gid = int(c.data.split(":")[2])
    await safe_edit(c.message, f"Игра ID {gid}: выберите действие.", reply_markup=finished_actions_kb(gid, admin=is_admin(c.from_user.id, c.from_user.username)))
    await safe_answer(c, )


@dp.callback_query(F.data.startswith("finished:result:"))
async def finished_result(c: CallbackQuery):
    metric_click(c.from_user.id)
    gid = int(c.data.split(":")[2])
    async with Session() as session:
        g = await get_game(session, gid)
        blue, red, vold = await get_team_rosters(session, gid)
        if not vold and getattr(g, "voldemort_id", None):
            vold = await session.get(Player, g.voldemort_id)
        b_avg = round(sum(p.rating for p in blue) / max(1, len(blue)), 1)
        r_avg = round(sum(p.rating for p in red) / max(1, len(red)), 1)
        fav = favorite_side(b_avg, r_avg)
        human = RESULT_HUMAN.get(getattr(g, "result_type", "") or "", "Исход не указан")
        blue_txt = "\n".join(f"- {full_name(p)} [{p.rating}]" for p in blue) or "—"
        red_txt  = "\n".join(f"- {full_name(p)} [{p.rating}]" for p in red)  or "—"
        notes = _get_notes(gid)

    notes_text = ('🖋️ Заметки:\n' + "\n".join('• ' + n['text'] for n in notes)) if notes else ''
    txt = (
        f"Игра ID {gid}: {getattr(g,'title','')}\n\n"
        f"🟦 Орден Феникса ({len(blue)}):\n{blue_txt}\n\n"
        f"🟪 Пожиратели + Воландеморт ({len(red)}):\n{red_txt}\n" + ("" if not vold else f"Воландеморт: {full_name(vold)} [{getattr(vold, 'rating', '—')}]\n") + "\n"
        f"Результат: {human}\n"
        f"Средний MMR — Орден Феникса: {b_avg}, Пожиратели: {r_avg}\n"
        f"Фаворит матча: {fav}\n"
        f"{notes_text}"
    )
    await safe_edit(c.message, txt, reply_markup=finished_actions_kb(gid, admin=is_admin(c.from_user.id, c.from_user.username)))
    await safe_answer(c, )

@dp.callback_query(F.data.startswith("finished:note:"))
async def finished_note(c: CallbackQuery, state: FSMContext):
    metric_click(c.from_user.id)
    if not is_admin(c.from_user.id, c.from_user.username):
        await safe_answer(c, "Только для админов.", show_alert=True); return
    gid = int(c.data.split(":")[2])
    await state.update_data(note_gid=gid)
    await state.set_state(CreateGameFSM.wait_note_text)
    await safe_edit(c.message, "Введите текст заметки одним сообщением:", reply_markup=finished_actions_kb(gid, admin=True))
    await safe_answer(c, )

@dp.message(CreateGameFSM.wait_note_text)
async def finished_note_text(m: Message, state: FSMContext):
    data = await state.get_data()
    gid = int(data.get("note_gid"))
    txt = (m.text or "").strip()
    if not txt:
        await m.answer("Пустая заметка не сохранена."); return
    _add_note(gid, m.from_user.id, txt)
    await state.clear()
    await m.answer("Заметка сохранена 🖋️")

# ===================== Applications (admin) =====================
@dp.callback_query(F.data == "admin:apps")
async def admin_apps(c: CallbackQuery):
    metric_click(c.from_user.id)
    if not is_admin(c.from_user.id, c.from_user.username):
        await safe_answer(c, "Только для админов.", show_alert=True); return
    apps = [a for a in _load_apps() if a.get("status") == "pending"]
    kb = InlineKeyboardBuilder()
    if not apps:
        kb.button(text="⬅️ Назад", callback_data="admin:menu")
        kb.adjust(1)
        await safe_edit(c.message, "Заявок пока нет.", reply_markup=kb.as_markup()); await safe_answer(c, ); return
    for a in apps:
        text = f"{a['name']} (user_id {a['user_id']})"
        kb.button(text=f"✅ Принять: {text}", callback_data=f"app:approve:{a['user_id']}")
        kb.button(text=f"❌ Отклонить: {text}", callback_data=f"app:reject:{a['user_id']}")
    kb.button(text="⬅️ Назад", callback_data="admin:menu")
    kb.adjust(1)
    await safe_edit(c.message, "Заявки в Бота:", reply_markup=kb.as_markup())
    await safe_answer(c, )

@dp.callback_query(F.data.startswith("app:approve:"))
async def app_approve(c: CallbackQuery):
    metric_click(c.from_user.id)
    if not is_admin(c.from_user.id, c.from_user.username):
        await safe_answer(c, "Только для админов.", show_alert=True); return
    uid = int(c.data.split(":")[2])
    apps = _load_apps()
    app = next((a for a in apps if a["user_id"] == uid and a["status"] == "pending"), None)
    if not app:
        await safe_answer(c, "Заявка не найдена.", show_alert=True); return
    parts = app["name"].split()
    first, last = parts[0], (" ".join(parts[1:]) if len(parts) > 1 else None)
    async with Session() as session:
        new_player = await create_player(session, first_name=first, last_name=last, username=app.get("tg_username"))
        link_user_to_player(uid, new_player.id)
    try:
        await bot.send_message(app["chat_id"], "Добрый день! Вам открыт доступ к боту! Хороших игр ❤️")
    except Exception:
        pass
    app["status"] = "approved"
    _save_apps(apps)
    metric_inc("auth_approved")
    await safe_answer(c, "Заявка принята.")
    await admin_menu(c, None)

@dp.callback_query(F.data.startswith("app:reject:"))
async def app_reject(c: CallbackQuery):
    metric_click(c.from_user.id)
    if not is_admin(c.from_user.id, c.from_user.username):
        await safe_answer(c, "Только для админов.", show_alert=True); return
    uid = int(c.data.split(":")[2])
    apps = _load_apps()
    app = next((a for a in apps if a["user_id"] == uid and a["status"] == "pending"), None)
    if not app:
        await safe_answer(c, "Заявка не найдена.", show_alert=True); return
    try:
        await bot.send_message(app["chat_id"], "Упс! Что-то пошло не так, проверьте правильность введённых данных и попробуйте авторизоваться ещё раз!")
    except Exception:
        pass
    app["status"] = "rejected"
    _save_apps(apps)
    await safe_answer(c, "Заявка отклонена.")
    await admin_menu(c, None)

# ===================== Admin utils =====================
@dp.callback_query(F.data == "admin:menu")
async def admin_menu(c: CallbackQuery, state: FSMContext):
    metric_click(c.from_user.id)
    if await _maybe_warn_unfinished(c, state, "admin:menu"):
        return
    if not is_admin(c.from_user.id, c.from_user.username):
        await safe_answer(c, "Только для админов.", show_alert=True)
        return
    await safe_edit(c.message, "🛠 Админ-панель", reply_markup=admin_menu_kb())
    await safe_answer(c, )

@dp.callback_query(F.data == "admin:players")
async def admin_players(c: CallbackQuery, state: FSMContext):
    metric_click(c.from_user.id)
    if not is_admin(c.from_user.id, c.from_user.username):
        await safe_answer(c, "Только для админов.", show_alert=True)
        return
    async with Session() as session:
        res = await session.execute(select(Player).order_by(Player.first_name.asc(), Player.last_name.asc()))
        players = list(res.scalars().all())
    if not players:
        await safe_edit(c.message, "Пока нет игроков.", reply_markup=admin_menu_kb()); await safe_answer(c, ); return
    kb = InlineKeyboardBuilder()
    for p in players:
        label = f"{full_name(p)} (ID {p.id}, {p.rating})"
        kb.button(text=f"✏️ {label}", callback_data=f"admin:player:edit:{p.id}")
        kb.button(text=f"🗑 {label}", callback_data=f"admin:player:del:{p.id}")
    kb.button(text="⬅️ Назад", callback_data="admin:menu")
    kb.adjust(1)
    await safe_edit(c.message, "Игроки (редактирование / удаление):", reply_markup=kb.as_markup())
    await safe_answer(c, )

@dp.callback_query(F.data.startswith("admin:player:edit:"))
async def admin_player_edit(c: CallbackQuery, state: FSMContext):
    metric_click(c.from_user.id)
    if not is_admin(c.from_user.id, c.from_user.username):
        await safe_answer(c, "Только для админов.", show_alert=True); return
    _, _, _, pid_s = c.data.split(":")
    pid = int(pid_s)
    await state.update_data(edit_player_id=pid)
    kb = InlineKeyboardBuilder()
    kb.button(text="⬅️ Назад", callback_data="admin:players")
    kb.adjust(1)
    await safe_edit(
        c.message,
        "Введите новое имя **или** имя и фамилию через пробел.\n_Примеры:_ `Иван` или `Иван Токунов`",
        parse_mode="Markdown",
        reply_markup=kb.as_markup(),
    )
    await state.set_state(AdminFSM.wait_new_fullname)
    await safe_answer(c, )

@dp.message(AdminFSM.wait_new_fullname)
async def admin_player_apply_name(m: Message, state: FSMContext):
    if not is_admin(m.from_user.id, m.from_user.username):
        await m.answer("Только для админов."); return
    raw = (m.text or "").strip()
    if not raw:
        await m.answer("Введите хотя бы имя."); return
    parts = [x for x in raw.split() if x]
    first = parts[0]
    last = " ".join(parts[1:]) if len(parts) > 1 else None
    data = await state.get_data()
    pid = data.get("edit_player_id")
    if not pid:
        await m.answer("Не найден контекст редактирования. Откройте список игроков ещё раз.")
        await state.clear(); return
    async with Session() as session:
        ok = await update_player_name(session, pid, first, last)
        res = await session.execute(select(Player).order_by(Player.first_name.asc(), Player.last_name.asc()))
        players = list(res.scalars().all())
    kb = InlineKeyboardBuilder()
    for p in players:
        label = f"{full_name(p)} (ID {p.id}, {p.rating})"
        kb.button(text=f"✏️ {label}", callback_data=f"admin:player:edit:{p.id}")
        kb.button(text=f"🗑 {label}", callback_data=f"admin:player:del:{p.id}")
    kb.button(text="⬅️ Назад", callback_data="admin:menu")
    kb.adjust(1)
    if ok:
        await m.answer(f"Готово. Новое имя: *{first}{(' ' + last) if last else ''}*.", parse_mode="Markdown", reply_markup=kb.as_markup())
    else:
        await m.answer("Игрок не найден.", reply_markup=kb.as_markup())
    await state.clear()

@dp.callback_query(F.data.startswith("admin:player:del:"))
async def admin_player_delete(c: CallbackQuery, state: FSMContext):
    metric_click(c.from_user.id)
    if not is_admin(c.from_user.id, c.from_user.username):
        await safe_answer(c, "Только для админов.", show_alert=True); return
    _, _, _, pid_s = c.data.split(":")
    pid = int(pid_s)
    async with Session() as session:
        removed, msg = await delete_player_if_no_games(session, pid)
        res = await session.execute(select(Player).order_by(Player.first_name.asc(), Player.last_name.asc()))
        players = list(res.scalars().all())
    kb = InlineKeyboardBuilder()
    for p in players:
        label = f"{full_name(p)} (ID {p.id}, {p.rating})"
        kb.button(text=f"✏️ {label}", callback_data=f"admin:player:edit:{p.id}")
        kb.button(text=f"🗑 {label}", callback_data=f"admin:player:del:{p.id}")
    kb.button(text="⬅️ Назад", callback_data="admin:menu")
    kb.adjust(1)
    await safe_answer(c, msg if msg else ("Игрок удалён." if removed else "Операция завершена."))
    await safe_edit(c.message, "Игроки (редактирование / удаление):", reply_markup=kb.as_markup())

@dp.callback_query(F.data == "admin:games")
async def admin_games(c: CallbackQuery, state: FSMContext):
    metric_click(c.from_user.id)
    if not is_admin(c.from_user.id, c.from_user.username):
        await safe_answer(c, "Только для админов.", show_alert=True); return
    async with Session() as session:
        games = await list_all_games(session)
    if not games:
        await safe_edit(c.message, "Игр ещё нет.", reply_markup=admin_menu_kb()); await safe_answer(c, ); return
    await safe_edit(c.message, "Игры (удаление — последние 50):", reply_markup=admin_games_kb(games))
    await safe_answer(c, )

@dp.callback_query(F.data.startswith("admin:game:del:"))
async def admin_game_delete(c: CallbackQuery, state: FSMContext):
    metric_click(c.from_user.id)
    if not is_admin(c.from_user.id, c.from_user.username):
        await safe_answer(c, "Только для админов.", show_alert=True); return
    _, _, _, gid_s = c.data.split(":")
    gid = int(gid_s)
    async with Session() as session:
        await delete_game(session, gid)
        games = await list_all_games(session)
    await safe_answer(c, f"Игра {gid} удалена (каскадно удалены её участники).")
    if games:
        await safe_edit(c.message, "Игры (удаление — последние 50):", reply_markup=admin_games_kb(games))
    else:
        await safe_edit(c.message, "Игр больше нет.", reply_markup=admin_menu_kb())


@dp.callback_query(F.data == "admin:recompute")
async def admin_recompute(c: CallbackQuery, state: FSMContext):
    metric_click(c.from_user.id)
    if not is_admin(c.from_user.id, c.from_user.username):
        await safe_answer(c, "Только для админов.", show_alert=True); return
    from services import recompute_win_counters
    async with Session() as session:
        summary = await recompute_all_ratings(session)
        await recompute_win_counters(session)
    await safe_edit(c.message, f"✅ Пересчёт завершён.\n{summary}", reply_markup=admin_menu_kb())

    # Автоэкспорт Excel с актуальными данными
    from openpyxl import Workbook
    from sqlalchemy import select
    import tempfile, os
    async with Session() as session2:
        res2 = await session2.execute(select(Player).order_by(Player.rating.desc(), Player.first_name.asc(), Player.last_name.asc()))
        players2 = list(res2.scalars().all())
    tmp = tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx")
    file_path = tmp.name; tmp.close()
    try:
        wb = Workbook(); ws = wb.active; ws.title = "Рейтинг"
        ws.append(["#", "Имя", "Фамилия", "MMR", "Победы Ордена", "Победы Пожирателей (вкл. Воландеморта)", "Директором избран Воландеморт", "Игрок отправил Воландеморта в Азкабан"])
        for i, p in enumerate(players2, start=1):
            blue = int(getattr(p, 'blue_wins', 0) or 0)
            red  = int(getattr(p, 'red_wins', 0) or 0)
            vold = int(getattr(p, 'vold_wins', 0) or 0)
            ws.append([i, p.first_name, (p.last_name or ""), int(p.rating), int(getattr(p, "blue_wins", 0) or 0), int(getattr(p, "red_wins", 0) or 0) + int(getattr(p, "vold_wins", 0) or 0), int(getattr(p, "social_vold", 0) or 0), int(getattr(p, "killer_points", 0) or 0)])
        wb.save(file_path)
        await c.message.answer_document(FSInputFile(file_path), caption="Экспорт рейтинга (Excel)")
    finally:
        try: os.remove(file_path)
        except Exception: pass

    await safe_answer(c, "Рейтинг пересчитан.")
@dp.callback_query(F.data == "admin:info")
async def admin_info(c: CallbackQuery, state: FSMContext):
    metric_click(c.from_user.id)
    if not is_admin(c.from_user.id, c.from_user.username):
        await safe_answer(c, "Только для админов.", show_alert=True); return
    txt = """<b>ℹ️ Инфо для админов</b>

<u>Главное меню</u>
• <b>➕ Добавить игру</b> — создать новую игру, в заголовке фиксируется время (МСК).
• <b>📚 Завершённые игры</b> — список игр за период, просмотр составов, результатов и заметок.
• <b>🧩 Игрок дня</b> — топ игроков сегодняшнего дня по метрикам MMR/соц-очки/средний рейтинг соперников.
• <b>🏆 Рейтинг игроков</b> — общий рейтинг (MMR), экспорт в Excel, разрезы.
• <b>📊 Моя статистика</b> — личные позиции в рейтингах.
• <b>❓ FAQ</b> — правила расчётов и терминология.
• <b>🛠 Админ-панель</b> — раздел управления данными и метриками.
• <b>⬅️ В главное меню</b> — вернуться на стартовый экран.

<u>Меню игры</u>
• <b>Команда Ордена Феникса</b> — выбрать игроков из «Списка дня» или из всех игроков.
• <b>Команда Пожирателей</b> — выбрать игроков из «Списка дня» или из всех игроков.
• <b>Кто был Воландемортом</b> — отметить Воландеморта (не должен быть в Ордене Феникса).
• <b>Проверить заполнение</b> — валидация составов и статуса.
• <b>Выберите победителя</b> — зафиксировать исход. Если «Воландеморт убит» — выбрать киллера.

<u>Админ-панель</u>
• <b>🧑‍🤝‍🧑 Игроки</b> — переименование, удаление (если нет участий в играх).
• <b>🎮 Игры (удаление)</b> — каскадное удаление игры и её участников.
• <b>🔁 Пересчитать рейтинг (все)</b> — сбросить MMR/соц‑очки и пересчитать все игры.
• <b>📋 Список дня</b> — быстрый список игроков для набора команд.
• <b>📈 Статистика бота</b> — счётчики, активные пользователи, экспорт в Excel.
• <b>⬅️ В главное меню</b> — вернуться на стартовый экран.
"""
    await safe_edit(c.message, txt, parse_mode="HTML", reply_markup=admin_menu_kb())
    await safe_answer(c, )

@dp.callback_query(F.data == "botstats:menu")
async def botstats_menu(c: CallbackQuery):
    metric_click(c.from_user.id)
    if not is_admin(c.from_user.id, c.from_user.username):
        await safe_answer(c, "Только для админов.", show_alert=True); return
    await safe_edit(c.message, "Статистика бота — выберите период:", reply_markup=botstats_menu_kb())
    await safe_answer(c, )

@dp.callback_query(F.data.in_(("botstats:week","botstats:month","botstats:all")))
async def botstats_show(c: CallbackQuery):
    metric_click(c.from_user.id)
    if not is_admin(c.from_user.id, c.from_user.username):
        await safe_answer(c, "Только для админов.", show_alert=True); return
    mode = c.data.split(":")[1]
    text, _ = _metrics_summary(mode)
    await safe_edit(c.message, text, parse_mode="HTML", reply_markup=botstats_menu_kb())
    await safe_answer(c, )

@dp.callback_query(F.data == "botstats:export")
async def botstats_export(c: CallbackQuery):
    metric_click(c.from_user.id)
    if not is_admin(c.from_user.id, c.from_user.username):
        await safe_answer(c, "Только для админов.", show_alert=True); return
    from openpyxl import Workbook
    m = _load_json_obj(METRICS_PATH)
    with tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx") as tmp:
        file_path = tmp.name
    try:
        wb = Workbook()
        ws = wb.active
        ws.title = "Статистика бота"
        ws.append(["Дата", "Уникальных пользователей", "Кликов (значимых)"])
        for day, obj in sorted(m.get("by_day", {}).items()):
            ws.append([day, len(set(obj.get("active_user_ids", []))), int(obj.get("clicks", 0))])
        ws2 = wb.create_sheet("Счётчики")
        ws2.append(["Метрика", "Значение"])
        for k, v in m.get("counters", {}).items():
            ws2.append([k, v])
        wb.save(file_path)
        await c.message.answer_document(FSInputFile(file_path), caption="Статистика бота (Excel)")
        await safe_answer(c, "Файл готов.")
    finally:
        try: os.remove(file_path)
        except Exception: pass

# ===================== Player of the Day =====================
@dp.callback_query(F.data == "playeroftheday")
async def player_of_the_day(c: CallbackQuery, state: FSMContext):
    metric_click(c.from_user.id)
    if await _maybe_warn_unfinished(c, state, "playeroftheday"):
        return
    stats_path = Path("game_stats.json")
    data = []
    try:
        with stats_path.open("r", encoding="utf-8") as f:
            data = json.load(f) or []
    except Exception:
        pass
    today = now_msk().date().isoformat()

    agg: Dict[int, Dict[str, float]] = {}
    for rec in data:
        ts = rec.get("ts")
        if not ts:
            continue
        try:
            if datetime.fromisoformat(ts).date().isoformat() != today:
                continue
        except Exception:
            continue
        pid = int(rec.get("player_id", 0))
        mmr = int(rec.get("mmr_delta", 0))
        soc = int(rec.get("social_gain", 0))
        opp = float(rec.get("opponent_avg", 0.0))
        a = agg.setdefault(pid, {"mmr": 0, "soc": 0, "opp_sum": 0.0, "cnt": 0})
        a["mmr"] += mmr
        a["soc"] += soc
        a["opp_sum"] += opp
        a["cnt"] += 1

    if not agg:
        admin = is_admin(c.from_user.id, c.from_user.username)
        await safe_edit(c.message, "За сегодня игр ещё не было.", reply_markup=home_kb_for_user(admin, is_authorized_user(c.from_user.id)))
        await safe_answer(c, ); return

    def key(pid):
        a = agg[pid]
        opp_avg = (a["opp_sum"] / a["cnt"]) if a["cnt"] else 0.0
        return (a["mmr"], a["soc"], opp_avg)

    async with Session() as session:
        res = await session.execute(select(Player).where(Player.id.in_(list(agg.keys()))))
        players = {p.id: p for p in res.scalars().all()}

    ranked = sorted(agg.keys(), key=key, reverse=True)
    top = ranked[:5]
    lines = []
    for i, pid in enumerate(top, 1):
        p = players.get(pid)
        if not p:
            continue
        a = agg[pid]
        opp_avg = (a["opp_sum"] / a["cnt"]) if a["cnt"] else 0.0
        lines.append(f"{i}. {full_name(p)} — MMR за день: {a['mmr']}, соц: {a['soc']}, Бухгольц: {opp_avg:.1f}")

    admin = is_admin(c.from_user.id, c.from_user.username)
    await safe_edit(
        c.message,
        "🧩 *Игрок дня*\n\n" + "\n".join(lines) + ("\n\nПобедитель определяется по MMR → соц-очкам → среднему рейтингу соперников (Бухгольц)." ),
        parse_mode="Markdown",
        reply_markup=home_kb_for_user(admin, is_authorized_user(c.from_user.id))
    )
    await safe_answer(c, )

# ===================== My stats =====================
@dp.callback_query(F.data == "me:stats")
async def my_stats(c: CallbackQuery, state: FSMContext):
    metric_click(c.from_user.id)
    if await _maybe_warn_unfinished(c, state, "me:stats"):
        return
    pid = get_player_id_for_user(c.from_user.id)
    if not pid:
        await safe_answer(c, "Вы не авторизованы.", show_alert=True); return

    async with Session() as session:
        me = await session.get(Player, pid)
        if not me:
            await safe_answer(c, "Игрок не найден. Обратитесь к администратору.", show_alert=True); return

        # --- Overall ranks (оставляем как было) ---
        res_all = await session.execute(select(Player).order_by(Player.rating.desc(), Player.first_name.asc(), Player.last_name.asc()))
        players = list(res_all.scalars().all())

        def rank_by(key):
            arr = sorted(players, key=key, reverse=True)
            for idx, p in enumerate(arr, 1):
                if p.id == me.id:
                    return idx, len(arr)
            return None, len(arr)

        r_mmr, total = rank_by(lambda p: p.rating)
        r_blue, _ = rank_by(lambda p: p.social_blue)
        r_red, _  = rank_by(lambda p: p.social_red)
        r_vold, _ = rank_by(lambda p: p.social_vold)
        r_kill, _ = rank_by(lambda p: p.killer_points)

        # --- Streaks ---
        s = await get_player_streaks(session, pid)

        # =======================
        # 1) Совместная статистика с игроками (за ВСЕ игры) — только игры, где были в ОДНОЙ команде.
        #    Пожиратели и Воландеморт считаются одной стороной ("red").
        # =======================
        from sqlalchemy import select as _select
        resg = await session.execute(_select(Game).where(Game.result_type.is_not(None)).order_by(Game.id.asc()))
        all_games = list(resg.scalars().all())

        co_stats_all = {}  # pid -> {'games': int, 'wins': int}
        co_ids_all = set()

        for g in all_games:
            # Получаем составы
            resp = await session.execute(_select(GameParticipant).where(GameParticipant.game_id == g.id))
            parts = list(resp.scalars().all())
            blue_ids = [gp.player_id for gp in parts if gp.team == 'blue']
            red_ids  = [gp.player_id for gp in parts if gp.team == 'red']
            vold_ids = [gp.player_id for gp in parts if gp.team == 'voldemort']
            vold_id  = g.voldemort_id or (vold_ids[0] if vold_ids else None)

            # Расширяем красную сторону Воландемортом
            red_ext = list(red_ids)
            if vold_id and vold_id not in red_ext:
                red_ext.append(vold_id)

            # Определяем сторону текущего пользователя в этой игре
            my_side = None
            if pid in blue_ids:
                my_side = 'blue'
            elif pid in red_ext:
                my_side = 'red'
            else:
                continue  # не участвовал

            winner = 'blue' if (g.result_type or '').startswith('blue_') else 'red'

            # Список тиммейтов (только одна сторона со мной)
            same_side_ids = blue_ids if my_side == 'blue' else red_ext
            for cid in same_side_ids:
                if cid == pid:
                    continue
                co_ids_all.add(cid)
                st = co_stats_all.get(cid, {'games': 0, 'wins': 0})
                st['games'] += 1
                if winner == my_side:
                    st['wins'] += 1
                co_stats_all[cid] = st

        # Разрешаем имена для ко-игроков
        co_names_all = {}
        if co_ids_all:
            resp2 = await session.execute(_select(Player).where(Player.id.in_(list(co_ids_all))))
            for p2 in resp2.scalars().all():
                nm = f"{p2.first_name}{(' ' + p2.last_name) if p2.last_name else ''}"
                co_names_all[p2.id] = nm

        def win_pct_all(pid_):
            st = co_stats_all.get(pid_, {'games': 0, 'wins': 0})
            if st['games'] == 0:
                return 0.0
            return (st['wins'] / st['games']) * 100.0

        def loss_pct_all(pid_):
            st = co_stats_all.get(pid_, {'games': 0, 'wins': 0})
            if st['games'] == 0:
                return 0.0
            losses = st['games'] - st['wins']
            return (losses / st['games']) * 100.0

        def sort_key_for_top(lst_fn_pct, pid_):
            # сортируем по проценту (убыв.), затем по совместным играм (убыв.), затем по имени (возр.)
            pct = lst_fn_pct(pid_)
            games_cnt = co_stats_all.get(pid_, {}).get('games', 0)
            name = co_names_all.get(pid_, "")
            return (-pct, -games_cnt, name)

        co_list_all = list(co_stats_all.keys())
        top_win_all  = sorted(co_list_all, key=lambda x: sort_key_for_top(win_pct_all, x))[:5]
        top_lose_all = sorted(co_list_all, key=lambda x: sort_key_for_top(loss_pct_all, x))[:5]

        def fmt_top_all(lst, pct_fn):
            if not lst:
                return "—"
            out = []
            for idx, pid2 in enumerate(lst, 1):
                name = co_names_all.get(pid2, f"ID {pid2}")
                st = co_stats_all.get(pid2, {'games': 0, 'wins': 0})
                out.append(f"{idx}. {name} — {pct_fn(pid2):.0f}% (совм. игр: {st['games']})")
            return "\n".join(out)

        top_win_block = fmt_top_all(top_win_all, win_pct_all)
        top_lose_block = fmt_top_all(top_lose_all, loss_pct_all)

        # =======================
        # 2) Блок "последние 10 игр" — переносим в самый низ.
        # =======================
        # собираем ID игр, где участвовал пользователь
        resp = await session.execute(_select(GameParticipant.game_id).where(GameParticipant.player_id == pid))
        gp_ids = set(resp.scalars().all())
        resv = await session.execute(_select(Game.id).where(Game.voldemort_id == pid))
        v_ids = set(resv.scalars().all())
        all_ids = list(gp_ids | v_ids)
        last_games = []
        if all_ids:
            resg10 = await session.execute(
                _select(Game).where(Game.id.in_(all_ids), Game.result_type.is_not(None)).order_by(Game.id.desc()).limit(10)
            )
            last_games = list(resg10.scalars().all())

        blue_wins = blue_losses = red_wins = red_losses = 0
        game_lines = []

        for g in last_games:
            parts_res = await session.execute(_select(GameParticipant).where(GameParticipant.game_id == g.id))
            parts = list(parts_res.scalars().all())
            blue_ids = [gp.player_id for gp in parts if gp.team == 'blue']
            red_ids  = [gp.player_id for gp in parts if gp.team == 'red']
            vold_ids = [gp.player_id for gp in parts if gp.team == 'voldemort']
            vold_id  = g.voldemort_id or (vold_ids[0] if vold_ids else None)

            red_ext = list(red_ids)
            if vold_id and vold_id not in red_ext:
                red_ext.append(vold_id)

            side = 'blue' if pid in blue_ids else ('red' if pid in red_ext else None)
            winner = 'blue' if (g.result_type or '').startswith('blue_') else 'red'

            if side == 'blue':
                if winner == 'blue': blue_wins += 1
                else: blue_losses += 1
            elif side == 'red':
                if winner == 'red': red_wins += 1
                else: red_losses += 1

            ts = getattr(g, "created_at", None)
            ts_str = ts.strftime("%d.%m.%Y %H:%M") if ts else getattr(g, "title", f"Игра {g.id}")
            side_h = "Орден" if side == 'blue' else ("Пожиратели" if side == 'red' else "—")
            outcome_h = "Победа" if side and winner == side else "Поражение"
            game_lines.append(f"• {ts_str} — ID {g.id} — {side_h} — {outcome_h}")

    admin = is_admin(c.from_user.id, c.from_user.username)

    # Формируем текст: сначала сводные показатели, затем «ТОПы», и в самый низ — последние 10 игр.
    n_last = len(game_lines)
    last_hdr = f"За последние {n_last} игр:" if n_last else "За последние 0 игр:"
    last_lines = "\n".join(game_lines)

    text = (
        f"📈 *Моя статистика*\n\n"
        f"MMR: *{me.rating}*  _(место: {r_mmr}/{total})_\n"
        f"Орден — соц очки: *{me.social_blue}*  _(место: {r_blue})_\n"
        f"Пожиратели — соц очки: *{me.social_red}*  _(место: {r_red})_\n"
        f"Воландеморт — соц очки: *{me.social_vold}*  _(место: {r_vold})_\n"
        f"Киллер Воландеморта — убийств: *{me.killer_points}*  _(место: {r_kill})_\n\n"
        f"Винстрики: winstreak *{s['cur_win']}* (макс: {s['max_win']}), losestreak *{s['cur_lose']}* (макс: {s['max_lose']})\n\n"
        f"*Наибольший процент побед с игроками (за все игры):*\n{top_win_block}\n\n"
        f"*Наибольший процент поражений с игроками (за все игры):*\n{top_lose_block}\n\n"
        f"*{last_hdr}*\n"
        f"{last_lines}\n\n"
        f"Побед за Орден Феникса: *{blue_wins}*, Поражений: *{blue_losses}*\n"
        f"Побед за Пожирателей: *{red_wins}*, Поражений: *{red_losses}*"
    )

    await safe_edit(
        c.message,
        text,
        parse_mode="Markdown",
        reply_markup=home_kb_for_user(admin, True),
    )
    await safe_answer(c, )


@dp.callback_query(F.data == "me:streak")
async def me_streak(c: CallbackQuery, state: FSMContext):
    metric_click(c.from_user.id)
    pid = get_player_id_for_user(c.from_user.id)
    if not pid:
        await safe_answer(c, "Вы не авторизованы.", show_alert=True); return

    async with Session() as session:
        s = await get_player_streaks(session, pid)

    text = (
        "📈 <b>Ваши стрики</b>\n\n"
        f"• Ваш winstreak: <b>{s['max_win']}</b>\n"
        f"• Ваш losestreak: <b>{s['max_lose']}</b>\n"
        f"• Ваш активный winstreak: <b>{s['cur_win']}</b>\n"
        f"• Ваш активный losestreak: <b>{s['cur_lose']}</b>"
    )
    await safe_edit(
        c.message,
        text,
        parse_mode="HTML",
        reply_markup=home_kb_for_user(is_admin(c.from_user.id, c.from_user.username), True),
    )
    await safe_answer(c, )

# ===================== Galleons / Shop Handlers =====================
@dp.callback_query(F.data == "me:galleons")
async def me_galleons(c: CallbackQuery, state: FSMContext):
    metric_click(c.from_user.id)
    # Не зависаем даже без привязки
    pid = get_player_id_for_user(c.from_user.id)
    galls = 0
    if pid:
        async with Session() as session:
            p = await session.get(Player, pid)
            if p and getattr(p, "galleons_balance", None) is not None:
                galls = int(p.galleons_balance)
    text = f"Количество Галлеонов {COIN} {galls}"
    await safe_edit(c.message, text, reply_markup=home_kb_for_user(is_admin(c.from_user.id, c.from_user.username), True))
    await safe_answer(c, )

@dp.callback_query(F.data == "shop:menu")
async def shop_menu(c: CallbackQuery, state: FSMContext):
    metric_click(c.from_user.id)
    if not is_authorized_user(c.from_user.id):
        await safe_answer(c, "Вы не авторизованы.", show_alert=True); return
    await safe_edit(c.message, "Лавка Олливандера. Выберите товар:", reply_markup=shop_menu_kb())
    await safe_answer(c, )

@dp.callback_query(F.data.startswith("shop:buy:"))
async def shop_buy(c: CallbackQuery, state: FSMContext):
    metric_click(c.from_user.id)
    if not is_authorized_user(c.from_user.id):
        await safe_answer(c, "Вы не авторизованы.", show_alert=True); return
    code = c.data.split(":", 2)[2]
    item = next((i for i in SHOP_ITEMS if i["code"] == code), None)
    if not item:
        await safe_answer(c, "Товар не найден.", show_alert=True); return
    pid = get_player_id_for_user(c.from_user.id)
    async with Session() as session:
        p = await session.get(Player, pid)
        balance = p.galleons_balance
    if balance < item["cost"]:
        await safe_answer(c, f"Недостаточно Галлеонов 🪙 🪙 для покупки «{item['title']}».", show_alert=True); return
    kb = InlineKeyboardBuilder()
    kb.button(text="Да", callback_data=f"shop:confirm:{code}")
    kb.button(text="Нет", callback_data="shop:cancel")
    kb.adjust(2)
    await safe_edit(c.message, f"Вы точно хотите приобрести «{item['title']}» за {item['cost']}{COIN}?", reply_markup=kb.as_markup())
    await safe_answer(c, )

@dp.callback_query(F.data == "shop:cancel")
async def shop_cancel(c: CallbackQuery, state: FSMContext):
    metric_click(c.from_user.id)
    await shop_menu(c, state)

@dp.callback_query(F.data.startswith("shop:confirm:"))
async def shop_confirm(c: CallbackQuery, state: FSMContext):
    metric_click(c.from_user.id)
    if not is_authorized_user(c.from_user.id):
        await safe_answer(c, "Вы не авторизованы.", show_alert=True); return
    code = c.data.split(":", 2)[2]
    item = next((i for i in SHOP_ITEMS if i["code"] == code), None)
    if not item:
        await safe_answer(c, "Товар не найден.", show_alert=True); return
    pid = get_player_id_for_user(c.from_user.id)
    async with Session() as session:
        p = await session.get(Player, pid)
        if p.galleons_balance < item["cost"]:
            await safe_answer(c, "Недостаточно Галлеонов 🪙 🪙.", show_alert=True); return
        # Create purchase and deduct balance
        title = item["title"]
        pur = await create_purchase(session, pid, code, title, item["cost"])
        p.galleons_balance -= item["cost"]
        await session.commit()
    receipt_text = None
    if code == "pm_first_game":
        receipt_text = "Вы приобрели сертификат на заявление себя первым министром в первой игре вечера (до раздачи ролей)."
    elif code == "pm_replace_lord":
        receipt_text = "Вы приобрели сертификат на заявление себя министром сместив прошлого лорда."
    elif code == "badge":
        receipt_text = "Фирменный значок. Покажите данное сообщение сотруднику Антикафе."
    elif code == "random_12_rooms":
        receipt_text = "Вы приобрели случайный сертификат 12 комнат. Покажите данное сообщение сотруднику Антикафе."
    elif code == "named_ballot":
        receipt_text = "Вы приобрели именную голосовалку. Покажите данное сообщение сотруднику Антикафе."
    else:
        receipt_text = f"Вы приобрели: {item['title']}"
    receipt = (
        f"{item['emoji']} {receipt_text}\n"
        f"Время покупки: { _msk_now_str() }\n"
        f"С вас списано: {item['cost']} Галлеонов 🪙 🪙."
    )
    await safe_edit(c.message, receipt, parse_mode="HTML", reply_markup=home_kb_for_user(is_admin(c.from_user.id, c.from_user.username), True))

    await safe_answer(c, "Покупка оформлена.")


@dp.callback_query(F.data == "mypur:menu")
async def mypur_menu(c: CallbackQuery, state: FSMContext):
    metric_click(c.from_user.id)
    if not is_authorized_user(c.from_user.id):
        await safe_answer(c, "Вы не авторизованы.", show_alert=True); return
    pid = get_player_id_for_user(c.from_user.id)
    async with Session() as session:
        purchases = await list_purchases(session, pid)
    if not purchases:
        await safe_edit(c.message, "Пока пусто. Здесь будут ваши покупки.", reply_markup=home_kb_for_user(is_admin(c.from_user.id, c.from_user.username), True))
        await safe_answer(c, ); return
    await safe_edit(c.message, "Мои покупки:", reply_markup=mypurchases_list_kb(purchases))
    await safe_answer(c, )

@dp.callback_query(F.data.startswith("mypur:item:"))
async def mypur_item(c: CallbackQuery, state: FSMContext):
    metric_click(c.from_user.id)
    pid = get_player_id_for_user(c.from_user.id)
    if not pid:
        await safe_answer(c, "Вы не авторизованы.", show_alert=True); return
    pur_id = int(c.data.split(":")[2])
    async with Session() as session:
        from db import Purchase
        pur = await session.get(Purchase, pur_id)
    if not pur or pur.player_id != pid:
        await safe_answer(c, "Покупка не найдена.", show_alert=True); return
    text = f"Покупка: {pur.title}\nСтатус: {'✅ Получено' if pur.is_received else '❌ Не получено'}"
    await safe_edit(c.message, text, reply_markup=purchase_status_kb(pur_id))
    await safe_answer(c, )

@dp.callback_query(F.data.startswith("mypur:set:"))
async def mypur_set(c: CallbackQuery, state: FSMContext):
    metric_click(c.from_user.id)
    pid = get_player_id_for_user(c.from_user.id)
    if not pid:
        await safe_answer(c, "Вы не авторизованы.", show_alert=True); return
    _, _, pur_id, received = c.data.split(":")
    pur_id = int(pur_id); received = received == "1"
    async with Session() as session:
        from db import Purchase
        pur = await session.get(Purchase, pur_id)
        if not pur or pur.player_id != pid:
            await safe_answer(c, "Покупка не найдена.", show_alert=True); return
        ok = await set_purchase_received(session, pur_id, received)
    await mypur_menu(c, state)
    await safe_answer(c, "Статус обновлён.")

# Admin: recompute galleons
@dp.callback_query(F.data == "admin:recompute_galleons")
async def admin_recompute_galleons(c: CallbackQuery, state: FSMContext):
    metric_click(c.from_user.id)
    if not is_admin(c.from_user.id, c.from_user.username):
        await safe_answer(c, "Только для админов.", show_alert=True); return
    async with Session() as session:
        summary = await recompute_all_galleons(session)
    await safe_edit(c.message, f"✅ Пересчёт Галлеонов завершён.\n{summary}", reply_markup=admin_menu_kb())
    await safe_answer(c, "Галлеоны пересчитаны.")


# ===================== Fallback =====================
@dp.message()
async def fallback_any(m: Message, state: FSMContext):
    data = await state.get_data()
    game_id = data.get("game_id")
    admin = is_admin(m.from_user.id, m.from_user.username)
    if game_id:
        async with Session() as session:
            summary, *_ = await roster_summary(session, game_id)
            g = await get_game(session, game_id)
        await m.answer(f"Игра: *{getattr(g,'title','Игра')}*.\n\n{summary}", parse_mode="Markdown", reply_markup=main_menu_kb(game_id))
        return
    await m.answer(
        "Главное меню.\nЭтот бот ведёт рейтинги игры «Тайный Воландеморт».",
        reply_markup=home_kb_for_user(admin, is_authorized_user(m.from_user.id)),
    )

# ===================== run =====================
async def main():
    await init_db()
    try:
        await dp.start_polling(bot, allowed_updates=dp.resolve_used_update_types())
    finally:
        await bot.session.close()

if __name__ == "__main__":
    try:
        asyncio.run(main())
    except (KeyboardInterrupt, SystemExit, asyncio.CancelledError):
        logging.info("Bot stopped by user.")
